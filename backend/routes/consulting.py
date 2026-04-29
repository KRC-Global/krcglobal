"""
GBMS - Consulting Projects Routes
해외기술용역 프로젝트 관리 API
"""
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    Workbook = None

try:
    import pandas as pd
except ImportError:
    pd = None
from werkzeug.utils import secure_filename
from models import db, ConsultingProject, ConsultingPersonnel, ActivityLog, ProfitabilityData, ProposalStatus, ProjectLifecycle, Eoi, Proposal, Contract, PerformanceRecord, TorRfp, normalize_date_dot
from routes.auth import token_required, permission_required
from sqlalchemy.orm import joinedload
import os
import re

consulting_bp = Blueprint('consulting', __name__)


def parse_project_date(date_str, is_end=False):
    """'YY-MM 또는 YYYY-MM-DD 형식의 날짜 문자열을 date 객체로 변환
    is_end=True이면 월 단위 날짜(YY-MM)를 해당 월의 말일로 처리"""
    import calendar
    from datetime import date as date_type
    if isinstance(date_str, date_type):
        return date_str
    s = str(date_str).replace("'", "").replace("\u2018", "").replace("\u2019", "").strip()
    # YYYY-MM-DD 형식 시도
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except ValueError:
        pass
    # YY-MM 또는 YY.MM 형식 시도
    parts = re.split(r'[.\-]', s)
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        yy = int(parts[0])
        mm = int(parts[1])
        year = 1900 + yy if yy >= 50 else 2000 + yy
        if 1900 <= year <= 2100 and 1 <= mm <= 12:
            if is_end:
                # 종료일은 해당 월의 말일로 처리
                last_day = calendar.monthrange(year, mm)[1]
                return datetime(year, mm, last_day).date()
            return datetime(year, mm, 1).date()
    return None


def calculate_progress(project):
    """프로젝트 진행률 계산 및 실제 상태 반환"""
    status = project.status or '시행중'

    # 이미 준공/완료 상태면 그대로 반환
    if status in ['준공', '완료']:
        return 100, status

    progress = 0
    try:
        start = project.start_date
        end = project.end_date

        if start and end:
            now = datetime.now().date()
            start_date = parse_project_date(start, is_end=False)
            end_date = parse_project_date(end, is_end=True)

            if start_date and end_date:
                if now >= end_date:
                    progress = 100
                elif now <= start_date:
                    progress = 0
                else:
                    total_days = (end_date - start_date).days
                    elapsed_days = (now - start_date).days
                    if total_days > 0:
                        # end_date 전이면 최대 99%로 제한 (반올림으로 100% 오판 방지)
                        progress = min(round((elapsed_days / total_days) * 100), 99)
    except Exception:
        pass

    # 진행률 100%이고 시행중/진행중이면 준공으로 처리
    if progress >= 100 and status in ['시행중', '진행중']:
        return 100, '준공'

    return progress, status


def validate_project_data(data, is_update=False):
    """프로젝트 데이터 유효성 검증"""
    errors = []

    # 필수 필드 검증 (생성 시에만)
    if not is_update:
        if not data.get('titleKr'):
            errors.append('국문사업명은 필수 입력 항목입니다.')
        if not data.get('country'):
            errors.append('국가는 필수 입력 항목입니다.')

    # 수주년도 검증
    if 'contractYear' in data and data['contractYear']:
        try:
            year = int(data['contractYear'])
            if year < 1972 or year > 2100:
                errors.append('수주년도는 1972년부터 2100년 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('수주년도는 숫자여야 합니다.')

    # 예산 검증
    if 'budget' in data and data['budget']:
        try:
            budget = float(data['budget'])
            if budget < 0:
                errors.append('예산은 0 이상이어야 합니다.')
        except (ValueError, TypeError):
            errors.append('예산은 숫자여야 합니다.')

    # 좌표 검증
    if 'latitude' in data and data['latitude']:
        try:
            lat = float(data['latitude'])
            if lat < -90 or lat > 90:
                errors.append('위도는 -90도에서 90도 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('위도는 숫자여야 합니다.')

    if 'longitude' in data and data['longitude']:
        try:
            lon = float(data['longitude'])
            if lon < -180 or lon > 180:
                errors.append('경도는 -180도에서 180도 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('경도는 숫자여야 합니다.')

    # 상태 검증
    if 'status' in data and data['status']:
        allowed_statuses = ['준공', '진행중', '시행중', '제안서제출', 'EOI제출']
        if data['status'] not in allowed_statuses:
            errors.append(f'상태는 {", ".join(allowed_statuses)} 중 하나여야 합니다.')

    return errors


@consulting_bp.route('', methods=['GET'])
@token_required
def get_consulting_projects(current_user):
    """Get all consulting projects with filters"""
    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    country = request.args.get('country')
    status = request.args.get('status')
    year = request.args.get('year', type=int)
    client = request.args.get('client')
    search = request.args.get('search')
    include_personnel = request.args.get('include_personnel', 'false').lower() == 'true'

    # Build query — joinedload로 creator N+1 쿼리 방지
    query = ConsultingProject.query.options(joinedload(ConsultingProject.creator))

    if country:
        query = query.filter(ConsultingProject.country == country)

    if client:
        query = query.filter(ConsultingProject.client.ilike(f'%{client}%'))

    if search:
        query = query.filter(
            db.or_(
                ConsultingProject.title_kr.ilike(f'%{search}%'),
                ConsultingProject.title_en.ilike(f'%{search}%'),
                ConsultingProject.country.ilike(f'%{search}%'),
                ConsultingProject.client.ilike(f'%{search}%')
            )
        )

    # Order by contract year (descending) and number
    query = query.order_by(
        ConsultingProject.contract_year.desc(),
        ConsultingProject.number.asc()
    )

    projects = query.all()

    # ===== 진행률 기반 자동 상태 업데이트 =====
    # 계산 결과를 캐시해서 이후 재사용 (이중 호출 방지)
    progress_cache = {}
    updated_count = 0
    for project in projects:
        prog, effective_status = calculate_progress(project)
        progress_cache[project.id] = prog
        if prog >= 100 and project.status in ['시행중', '진행중']:
            project.status = '준공'
            project.updated_at = datetime.utcnow()
            updated_count += 1

    if updated_count > 0:
        db.session.commit()

    # Year filtering (Python-side due to custom date format)
    if year:
        filtered_projects = []
        for project in projects:
            # Check contract_year first (legacy behavior)
            if project.contract_year == year:
                filtered_projects.append(project)
                continue

            # Check duration (start_date ~ end_date)
            # Format: 'YY-MM, YY.MM, etc.
            try:
                start_year = None
                end_year = None

                def parse_year(date_str):
                    if not date_str:
                        return None
                    # Normalize string: remove quotes and spaces
                    clean_str = date_str.replace("'", "").replace("'", "").replace("'", "").strip()
                    # Split by common separators
                    import re
                    parts = re.split(r'[.\-~]', clean_str)
                    if parts and parts[0].isdigit():
                        yy = int(parts[0])
                        # Handle 2-digit years
                        if 0 <= yy <= 99:
                            return 1900 + yy if yy >= 50 else 2000 + yy
                        # Handle 4-digit years
                        if 1900 <= yy <= 2100:
                            return yy
                    return None

                start_year = parse_year(project.start_date)
                end_year = parse_year(project.end_date)

                # Logic:
                # If both: match if start_year <= year <= end_year
                # If only start_year: match if year >= start_year
                # If only end_year: match if year <= end_year

                if start_year and end_year:
                    if start_year <= year <= end_year:
                        filtered_projects.append(project)
                elif start_year:
                    if start_year <= year:
                        filtered_projects.append(project)
                elif end_year:
                    if year <= end_year:
                        filtered_projects.append(project)

            except Exception:
                continue

        projects = filtered_projects

    # Status filtering (이제 업데이트된 status 기준으로 필터링)
    if status:
        filtered_by_status = []
        for project in projects:
            if project.status == status:
                filtered_by_status.append(project)
        projects = filtered_by_status

    # Manual Pagination
    total = len(projects)
    import math
    pages = math.ceil(total / per_page) if total > 0 else 1
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_items = projects[start_idx:end_idx]

    # Convert to dict — progress는 캐시에서 재사용 (이중 계산 방지)
    data_with_progress = []
    for project in paginated_items:
        project_dict = project.to_dict()
        project_dict['progress'] = progress_cache.get(project.id, 0)

        if include_personnel:
            project_dict['personnel'] = [person.to_dict() for person in project.personnel]

        data_with_progress.append(project_dict)

    return jsonify({
        'success': True,
        'data': data_with_progress,
        'total': total,
        'pages': pages,
        'currentPage': page
    })


@consulting_bp.route('/<int:project_id>', methods=['GET'])
@token_required
def get_consulting_project(current_user, project_id):
    """Get single consulting project"""
    project = ConsultingProject.query.get_or_404(project_id)
    include_personnel = request.args.get('include_personnel', 'false').lower() == 'true'

    project_dict = project.to_dict()

    # 인력 정보 포함
    if include_personnel:
        project_dict['personnel'] = [person.to_dict() for person in project.personnel]

    return jsonify({
        'success': True,
        'data': project_dict
    })


@consulting_bp.route('', methods=['POST'])
@token_required
def create_consulting_project(current_user):
    """Create new consulting project"""
    data = request.get_json()

    if not data:
        return jsonify({
            'success': False,
            'message': '요청 데이터가 없습니다.'
        }), 400

    # 데이터 유효성 검증
    errors = validate_project_data(data, is_update=False)
    if errors:
        return jsonify({
            'success': False,
            'message': '입력 데이터 검증 실패',
            'errors': errors
        }), 400

    # 중복 체크 (동일한 국가, 사업명, 수주년도)
    if data.get('titleKr') and data.get('country'):
        existing = ConsultingProject.query.filter_by(
            title_kr=data['titleKr'].strip(),
            country=data['country'].strip()
        )
        if data.get('contractYear'):
            existing = existing.filter_by(contract_year=data['contractYear'])

        if existing.first():
            return jsonify({
                'success': False,
                'message': '동일한 국가, 사업명, 수주년도를 가진 프로젝트가 이미 존재합니다.'
            }), 409

    try:
        # Create new project
        project = ConsultingProject(
            number=data.get('number'),
            contract_year=data.get('contractYear'),
            status=data.get('status', '준공'),
            country=data.get('country', '').strip(),
            latitude=data.get('latitude'),
            longitude=data.get('longitude'),
            title_en=data.get('titleEn', '').strip() if data.get('titleEn') else None,
            title_kr=data.get('titleKr', '').strip(),
            project_type=data.get('projectType', '').strip() if data.get('projectType') else None,
            type_feasibility=data.get('typeFeasibility', False),
            type_masterplan=data.get('typeMasterplan', False),
            type_basic_design=data.get('typeBasicDesign', False),
            type_detailed_design=data.get('typeDetailedDesign', False),
            type_construction=data.get('typeConstruction', False),
            type_pmc=data.get('typePmc', False),
            project_type_etc=data.get('projectTypeEtc', '').strip() if data.get('projectTypeEtc') else None,
            start_date=normalize_date_dot(data.get('startDate')),
            end_date=normalize_date_dot(data.get('endDate')),
            budget=data.get('budget'),
            krc_budget=data.get('krcBudget'),
            krc_share_ratio=data.get('krcShareRatio'),
            client=data.get('client', '').strip() if data.get('client') else None,
            funding_source=data.get('fundingSource', '').strip() if data.get('fundingSource') else None,
            total_budget=data.get('totalBudget'),
            budget_usd=data.get('budgetUsd'),
            krc_budget_usd=data.get('krcBudgetUsd'),
            description=data.get('description', '').strip() if data.get('description') else None,
            description_en=data.get('descriptionEn', '').strip() if data.get('descriptionEn') else None,
            lead_company=data.get('leadCompany', '').strip() if data.get('leadCompany') else None,
            lead_company_ratio=data.get('leadCompanyRatio'),
            jv1=data.get('jv1', '').strip() if data.get('jv1') else None,
            jv1_ratio=data.get('jv1Ratio'),
            jv2=data.get('jv2', '').strip() if data.get('jv2') else None,
            jv2_ratio=data.get('jv2Ratio'),
            jv3=data.get('jv3', '').strip() if data.get('jv3') else None,
            jv3_ratio=data.get('jv3Ratio'),
            jv4=data.get('jv4', '').strip() if data.get('jv4') else None,
            jv4_ratio=data.get('jv4Ratio'),
            jv5=data.get('jv5', '').strip() if data.get('jv5') else None,
            jv5_ratio=data.get('jv5Ratio'),
            created_by=current_user.id
        )

        db.session.add(project)
        db.session.flush()  # Get the project ID

        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='create',
            entity_type='consulting_project',
            entity_id=project.id,
            description=f'{current_user.name}님이 해외기술용역 프로젝트를 생성했습니다: {project.title_kr}',
            ip_address=request.remote_addr
        )
        db.session.add(log)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': '프로젝트가 생성되었습니다.',
            'data': project.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'프로젝트 생성 중 오류가 발생했습니다: {str(e)}'
        }), 500


@consulting_bp.route('/<int:project_id>', methods=['PUT'])
@token_required
def update_consulting_project(current_user, project_id):
    """Update consulting project"""
    project = ConsultingProject.query.get_or_404(project_id)
    data = request.get_json()

    if not data:
        return jsonify({
            'success': False,
            'message': '수정할 데이터가 없습니다.'
        }), 400

    # 데이터 유효성 검증
    errors = validate_project_data(data, is_update=True)
    if errors:
        return jsonify({
            'success': False,
            'message': '입력 데이터 검증 실패',
            'errors': errors
        }), 400

    # 중복 체크 (자기 자신은 제외)
    # contractYear가 없으면 기존 프로젝트 값을 사용 — 연도 미포함 시 너무 넓은 검사로 409 오발생 방지
    if data.get('titleKr') and data.get('country'):
        check_year = data.get('contractYear') or project.contract_year
        existing = ConsultingProject.query.filter(
            ConsultingProject.id != project_id,
            ConsultingProject.title_kr == data['titleKr'].strip(),
            ConsultingProject.country == data['country'].strip(),
            ConsultingProject.contract_year == check_year
        )

        if existing.first():
            return jsonify({
                'success': False,
                'message': '동일한 국가, 사업명, 수주년도를 가진 프로젝트가 이미 존재합니다.'
            }), 409

    try:
        # Update fields
        if 'number' in data:
            project.number = data['number']
        if 'contractYear' in data:
            project.contract_year = data['contractYear']
        if 'status' in data:
            project.status = data['status']
        if 'country' in data:
            project.country = data['country'].strip() if data['country'] else None
        if 'latitude' in data:
            project.latitude = data['latitude'] if data['latitude'] else None
        if 'longitude' in data:
            project.longitude = data['longitude'] if data['longitude'] else None
        if 'titleEn' in data:
            project.title_en = data['titleEn'].strip() if data['titleEn'] else None
        if 'titleKr' in data:
            project.title_kr = data['titleKr'].strip()
        if 'projectType' in data:
            project.project_type = data['projectType'].strip() if data['projectType'] else None
        if 'typeFeasibility' in data:
            project.type_feasibility = data['typeFeasibility']
        if 'typeMasterplan' in data:
            project.type_masterplan = data['typeMasterplan']
        if 'typeBasicDesign' in data:
            project.type_basic_design = data['typeBasicDesign']
        if 'typeDetailedDesign' in data:
            project.type_detailed_design = data['typeDetailedDesign']
        if 'typeConstruction' in data:
            project.type_construction = data['typeConstruction']
        if 'typePmc' in data:
            project.type_pmc = data['typePmc']
        if 'projectTypeEtc' in data:
            project.project_type_etc = data['projectTypeEtc'].strip() if data['projectTypeEtc'] else None
        if 'startDate' in data:
            project.start_date = normalize_date_dot(data['startDate'])
        if 'endDate' in data:
            project.end_date = normalize_date_dot(data['endDate'])
        if 'budget' in data:
            project.budget = data['budget'] if data['budget'] else None
        if 'krcBudget' in data:
            project.krc_budget = data['krcBudget'] if data['krcBudget'] else None
        if 'krcShareRatio' in data:
            project.krc_share_ratio = data['krcShareRatio'] if data['krcShareRatio'] else None
        if 'client' in data:
            project.client = data['client'].strip() if data['client'] else None
        if 'fundingSource' in data:
            project.funding_source = data['fundingSource'].strip() if data['fundingSource'] else None
        if 'totalBudget' in data:
            project.total_budget = data['totalBudget'] if data['totalBudget'] else None
        if 'budgetUsd' in data:
            project.budget_usd = data['budgetUsd'] if data['budgetUsd'] else None
        if 'krcBudgetUsd' in data:
            project.krc_budget_usd = data['krcBudgetUsd'] if data['krcBudgetUsd'] else None
        if 'description' in data:
            project.description = data['description'].strip() if data['description'] else None
        if 'descriptionEn' in data:
            project.description_en = data['descriptionEn'].strip() if data['descriptionEn'] else None
        if 'leadCompany' in data:
            project.lead_company = data['leadCompany'].strip() if data['leadCompany'] else None
        if 'leadCompanyRatio' in data:
            project.lead_company_ratio = data['leadCompanyRatio'] if data['leadCompanyRatio'] else None
        if 'jv1' in data:
            project.jv1 = data['jv1'].strip() if data['jv1'] else None
        if 'jv1Ratio' in data:
            project.jv1_ratio = data['jv1Ratio'] if data['jv1Ratio'] else None
        if 'jv2' in data:
            project.jv2 = data['jv2'].strip() if data['jv2'] else None
        if 'jv2Ratio' in data:
            project.jv2_ratio = data['jv2Ratio'] if data['jv2Ratio'] else None
        if 'jv3' in data:
            project.jv3 = data['jv3'].strip() if data['jv3'] else None
        if 'jv3Ratio' in data:
            project.jv3_ratio = data['jv3Ratio'] if data['jv3Ratio'] else None
        if 'jv4' in data:
            project.jv4 = data['jv4'].strip() if data['jv4'] else None
        if 'jv4Ratio' in data:
            project.jv4_ratio = data['jv4Ratio'] if data['jv4Ratio'] else None
        if 'jv5' in data:
            project.jv5 = data['jv5'].strip() if data['jv5'] else None
        if 'jv5Ratio' in data:
            project.jv5_ratio = data['jv5Ratio'] if data['jv5Ratio'] else None

        project.updated_at = datetime.utcnow()

        # Log activity
        log = ActivityLog(
            user_id=current_user.id,
            action='update',
            entity_type='consulting_project',
            entity_id=project.id,
            description=f'{current_user.name}님이 해외기술용역 프로젝트를 수정했습니다: {project.title_kr}',
            ip_address=request.remote_addr
        )
        db.session.add(log)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': '프로젝트가 수정되었습니다.',
            'data': project.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'프로젝트 수정 중 오류가 발생했습니다: {str(e)}'
        }), 500


@consulting_bp.route('/<int:project_id>', methods=['DELETE'])
@token_required
def delete_consulting_project(current_user, project_id):
    """Delete consulting project"""
    project = ConsultingProject.query.get_or_404(project_id)

    project_title = project.title_kr

    # Log activity before deletion
    log = ActivityLog(
        user_id=current_user.id,
        action='delete',
        entity_type='consulting_project',
        entity_id=project.id,
        description=f'{current_user.name}님이 해외기술용역 프로젝트를 삭제했습니다: {project_title}',
        ip_address=request.remote_addr
    )
    db.session.add(log)

    # Delete related logs manually if necessary (though usually logs are kept or set to null)
    # But if there's a strict constraint or just to be safe:
    try:
        ActivityLog.query.filter_by(entity_type='consulting_project', entity_id=project.id).delete()
    except:
        pass # Ignore if fails

    db.session.delete(project)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': '프로젝트가 삭제되었습니다.'
    })


@consulting_bp.route('/export', methods=['GET'])
@token_required
def export_consulting_projects(current_user):
    """Export consulting projects to Excel"""
    if Workbook is None:
        return jsonify({'success': False, 'message': 'Excel 라이브러리(openpyxl)가 설치되지 않았습니다.'}), 500
    # Get filters from query parameters
    country = request.args.get('country')
    status = request.args.get('status')
    year = request.args.get('year', type=int)
    client = request.args.get('client')
    search = request.args.get('search')

    # Build query
    query = ConsultingProject.query.options(joinedload(ConsultingProject.creator))

    if country:
        query = query.filter(ConsultingProject.country == country)
    if status:
        query = query.filter(ConsultingProject.status == status)
    if client:
        query = query.filter(ConsultingProject.client.ilike(f'%{client}%'))
    if search:
        query = query.filter(
            db.or_(
                ConsultingProject.title_kr.ilike(f'%{search}%'),
                ConsultingProject.title_en.ilike(f'%{search}%'),
                ConsultingProject.country.ilike(f'%{search}%'),
                ConsultingProject.client.ilike(f'%{search}%')
            )
        )

    # Order by contract year and number
    query = query.order_by(
        ConsultingProject.contract_year.asc(),
        ConsultingProject.number.asc()
    )

    projects = query.all()
    
    # Year filtering (Python-side)
    if year:
        filtered_projects = []
        for project in projects:
            # Check contract_year first (legacy behavior)
            if project.contract_year == year:
                filtered_projects.append(project)
                continue
                
            # Check duration
            try:
                start_year = None
                end_year = None
                
                def parse_year(date_str):
                    if not date_str:
                        return None
                    # Normalize string: remove quotes and spaces
                    clean_str = date_str.replace("'", "").replace("‘", "").replace("’", "").strip()
                    # Split by common separators
                    import re
                    parts = re.split(r'[.\-~]', clean_str)
                    if parts and parts[0].isdigit():
                        yy = int(parts[0])
                        # Handle 2-digit years
                        if 0 <= yy <= 99:
                            return 1900 + yy if yy >= 50 else 2000 + yy
                        # Handle 4-digit years
                        if 1900 <= yy <= 2100:
                            return yy
                    return None
                
                start_year = parse_year(project.start_date)
                end_year = parse_year(project.end_date)
                
                if start_year and end_year:
                    if start_year <= year <= end_year:
                        filtered_projects.append(project)
                elif start_year:
                    if start_year <= year:
                        filtered_projects.append(project)
                elif end_year:
                    if year <= end_year:
                        filtered_projects.append(project)
            except Exception:
                continue
        projects = filtered_projects

    # Load lifecycle data for all projects
    project_ids = [p.id for p in projects]
    lifecycles = ProjectLifecycle.query.filter(
        ProjectLifecycle.consulting_project_id.in_(project_ids)
    ).all() if project_ids else []
    lifecycle_map = {lc.consulting_project_id: lc for lc in lifecycles}

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "해외기술컨설팅"

    # Define headers
    headers = [
        '번호', '수주년도', '진행여부', '국가별', 'X', 'Y',
        '영문사업명', '국문사업명',
        '사업형태', '타당성조사', '기본계획', '기본설계', '실시설계', '시공감리', 'PMC', '사업형태기타',
        '착수일', '준공일',
        '전체사업비(백만USD)', '전체용역비(백만원)', '전체용역비(천USD)',
        '공사지분율', '공사지분용역비(백만원)', '공사지분용역비(천USD)',
        '발주처', '재원',
        '주관사', '주관사지분율', 'JV1', 'JV1지분율', 'JV2', 'JV2지분율', 'JV3', 'JV3지분율', 'JV4', 'JV4지분율', 'JV5', 'JV5지분율',
        'EOI일', 'EOI완료', '제안서일', '제안서완료', '계약일', '계약완료',
        '착수(킥오프)일', '착수완료', '준공일(라이프사이클)', '준공완료'
    ]

    # Style for headers
    header_fill = PatternFill(start_color="0A3D62", end_color="0A3D62", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, project in enumerate(projects, 2):
        lc = lifecycle_map.get(project.id)
        col = 1
        ws.cell(row=row_num, column=col, value=project.number); col += 1
        ws.cell(row=row_num, column=col, value=project.contract_year); col += 1
        ws.cell(row=row_num, column=col, value=project.status); col += 1
        ws.cell(row=row_num, column=col, value=project.country); col += 1
        ws.cell(row=row_num, column=col, value=float(project.longitude) if project.longitude else None); col += 1
        ws.cell(row=row_num, column=col, value=float(project.latitude) if project.latitude else None); col += 1
        ws.cell(row=row_num, column=col, value=project.title_en); col += 1
        ws.cell(row=row_num, column=col, value=project.title_kr); col += 1
        # 사업형태 세부
        ws.cell(row=row_num, column=col, value=project.project_type); col += 1
        ws.cell(row=row_num, column=col, value='O' if project.type_feasibility else ''); col += 1
        ws.cell(row=row_num, column=col, value='O' if project.type_masterplan else ''); col += 1
        ws.cell(row=row_num, column=col, value='O' if project.type_basic_design else ''); col += 1
        ws.cell(row=row_num, column=col, value='O' if project.type_detailed_design else ''); col += 1
        ws.cell(row=row_num, column=col, value='O' if project.type_construction else ''); col += 1
        ws.cell(row=row_num, column=col, value='O' if project.type_pmc else ''); col += 1
        ws.cell(row=row_num, column=col, value=project.project_type_etc or ''); col += 1
        # 착수일/준공일
        ws.cell(row=row_num, column=col, value=project.start_date); col += 1
        ws.cell(row=row_num, column=col, value=project.end_date); col += 1
        # 예산
        ws.cell(row=row_num, column=col, value=float(project.total_budget) if project.total_budget else None); col += 1
        ws.cell(row=row_num, column=col, value=float(project.budget) if project.budget else None); col += 1
        ws.cell(row=row_num, column=col, value=float(project.budget_usd) if project.budget_usd else None); col += 1
        ws.cell(row=row_num, column=col, value=float(project.krc_share_ratio) if project.krc_share_ratio else None); col += 1
        ws.cell(row=row_num, column=col, value=float(project.krc_budget) if project.krc_budget else None); col += 1
        ws.cell(row=row_num, column=col, value=float(project.krc_budget_usd) if project.krc_budget_usd else None); col += 1
        # 발주처/재원
        ws.cell(row=row_num, column=col, value=project.client); col += 1
        ws.cell(row=row_num, column=col, value=project.funding_source); col += 1
        # 컨소시엄 (회사명 + 지분율)
        ws.cell(row=row_num, column=col, value=project.lead_company); col += 1
        ws.cell(row=row_num, column=col, value=float(project.lead_company_ratio) if project.lead_company_ratio else None); col += 1
        for jv_idx in range(1, 6):
            ws.cell(row=row_num, column=col, value=getattr(project, f'jv{jv_idx}', None)); col += 1
            ratio = getattr(project, f'jv{jv_idx}_ratio', None)
            ws.cell(row=row_num, column=col, value=float(ratio) if ratio else None); col += 1
        # 라이프사이클 일자 + 완료여부
        ws.cell(row=row_num, column=col, value=lc.eoi_date if lc else None); col += 1
        ws.cell(row=row_num, column=col, value='O' if lc and lc.eoi_completed else ''); col += 1
        ws.cell(row=row_num, column=col, value=lc.proposal_date if lc else None); col += 1
        ws.cell(row=row_num, column=col, value='O' if lc and lc.proposal_completed else ''); col += 1
        ws.cell(row=row_num, column=col, value=lc.contract_date if lc else None); col += 1
        ws.cell(row=row_num, column=col, value='O' if lc and lc.contract_completed else ''); col += 1
        ws.cell(row=row_num, column=col, value=lc.kickoff_date if lc else None); col += 1
        ws.cell(row=row_num, column=col, value='O' if lc and lc.kickoff_completed else ''); col += 1
        ws.cell(row=row_num, column=col, value=lc.completion_date if lc else None); col += 1
        ws.cell(row=row_num, column=col, value='O' if lc and lc.completion_completed else ''); col += 1

    # Adjust column widths
    column_widths = [
        8, 10, 10, 15, 12, 12, 35, 35,                          # 번호~국문사업명
        20, 10, 10, 10, 10, 10, 10, 15,                          # 사업형태 세부
        12, 12,                                                    # 착수일/준공일
        18, 18, 18, 12, 18, 18,                                   # 예산
        20, 15,                                                    # 발주처/재원
        15, 10, 15, 10, 15, 10, 15, 10, 15, 10, 15, 10,          # 컨소시엄
        12, 8, 12, 8, 12, 8, 12, 8, 12, 8                        # 라이프사이클
    ]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    filename = f"해외기술용역_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # Log activity (실패해도 파일 전송은 계속)
    try:
        log = ActivityLog(
            user_id=current_user.id,
            action='export',
            entity_type='consulting_project',
            description=f'{current_user.name}님이 해외기술용역 프로젝트 {len(projects)}건을 Excel로 다운로드했습니다.',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()
    except Exception as log_err:
        print(f'[export] ActivityLog 저장 실패 (파일 전송 계속): {log_err}')
        db.session.rollback()

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@consulting_bp.route('/stats', methods=['GET'])
@token_required
def get_consulting_stats(current_user):
    """Get consulting projects statistics with filters"""
    # Get query parameters
    country = request.args.get('country')
    status = request.args.get('status')
    year = request.args.get('year', type=int)
    client = request.args.get('client')
    search = request.args.get('search')

    # Build query
    query = ConsultingProject.query

    if country:
        query = query.filter(ConsultingProject.country == country)
    if status:
        query = query.filter(ConsultingProject.status == status)
    if client:
        query = query.filter(ConsultingProject.client.ilike(f'%{client}%'))
    if search:
        query = query.filter(
            db.or_(
                ConsultingProject.title_kr.ilike(f'%{search}%'),
                ConsultingProject.title_en.ilike(f'%{search}%'),
                ConsultingProject.country.ilike(f'%{search}%'),
                ConsultingProject.client.ilike(f'%{search}%')
            )
        )

    projects = query.all()

    # Year filtering (Python-side)
    if year:
        filtered_projects = []
        for project in projects:
            # Check contract_year first (legacy behavior)
            if project.contract_year == year:
                filtered_projects.append(project)
                continue
            
            try:
                start_year = None
                end_year = None
                
                def parse_year(date_str):
                    if not date_str:
                        return None
                    clean_str = date_str.replace("'", "").replace("‘", "").replace("’", "").strip()
                    import re
                    parts = re.split(r'[.\-~]', clean_str)
                    if parts and parts[0].isdigit():
                        yy = int(parts[0])
                        if 0 <= yy <= 99:
                            return 1900 + yy if yy >= 50 else 2000 + yy
                        if 1900 <= yy <= 2100:
                            return yy
                    return None
                
                start_year = parse_year(project.start_date)
                end_year = parse_year(project.end_date)
                
                if start_year and end_year:
                    if start_year <= year <= end_year:
                        filtered_projects.append(project)
                elif start_year:
                    if start_year <= year:
                        filtered_projects.append(project)
                elif end_year:
                    if year <= end_year:
                        filtered_projects.append(project)
            except Exception:
                continue
        projects = filtered_projects

    # Calculate statistics from filtered projects
    total = len(projects)
    total_budget = sum(float(p.budget) for p in projects if p.budget)

    # Status counts (진행률 기반 실제 상태로 계산)
    status_counts = {}
    in_progress_budget = 0
    for p in projects:
        _, effective_status = calculate_progress(p)
        status_counts[effective_status] = status_counts.get(effective_status, 0) + 1
        # 시행중 사업의 용역비 합계 (금년도 용역비)
        if effective_status == '시행중' and p.budget:
            in_progress_budget += float(p.budget)

    # Country counts (top 10)
    country_counts_map = {}
    for p in projects:
        c = p.country or '미지정'
        country_counts_map[c] = country_counts_map.get(c, 0) + 1
    
    sorted_countries = sorted(country_counts_map.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Year counts (from filtered data)
    year_counts_map = {}
    for p in projects:
        y = p.contract_year or 0
        if y not in year_counts_map:
            year_counts_map[y] = {'count': 0, 'budget': 0}
        year_counts_map[y]['count'] += 1
        year_counts_map[y]['budget'] += float(p.budget) if p.budget else 0
        
    sorted_years = sorted(
        [{'year': y, 'count': d['count'], 'budget': d['budget']} for y, d in year_counts_map.items()],
        key=lambda x: x['year'], 
        reverse=True
    )[:10]

    # Projects with coordinates
    with_coords = sum(1 for p in projects if p.latitude and p.longitude)

    return jsonify({
        'success': True,
        'data': {
            'total': total,
            'byStatus': status_counts,
            'byCountry': [{'country': c, 'count': n} for c, n in sorted_countries],
            'byYear': sorted_years,
            'totalBudget': total_budget,
            'inProgressBudget': in_progress_budget,
            'withCoordinates': with_coords,
            'coordinateRate': round(with_coords / total * 100, 1) if total > 0 else 0
        }
    })


@consulting_bp.route('/countries', methods=['GET'])
@token_required
def get_consulting_countries(current_user):
    """Get list of countries with projects"""
    countries = db.session.query(
        ConsultingProject.country,
        db.func.count(ConsultingProject.id).label('count')
    ).group_by(ConsultingProject.country).order_by(
        ConsultingProject.country.asc()
    ).all()

    return jsonify({
        'success': True,
        'data': [
            {'country': country, 'count': count}
            for country, count in countries
        ]
    })


@consulting_bp.route('/clients', methods=['GET'])
@token_required
def get_consulting_clients(current_user):
    """Get list of clients"""
    clients = db.session.query(
        ConsultingProject.client
    ).filter(
        ConsultingProject.client.isnot(None)
    ).distinct().order_by(
        ConsultingProject.client.asc()
    ).all()

    return jsonify({
        'success': True,
        'data': [client[0] for client in clients if client[0]]
    })


@consulting_bp.route('/upload', methods=['POST'])
@token_required
def upload_consulting_projects(current_user):
    """Excel 파일을 통한 프로젝트 일괄 업로드"""

    if 'file' not in request.files:
        return jsonify({
            'success': False,
            'message': '파일이 전송되지 않았습니다.'
        }), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({
            'success': False,
            'message': '파일이 선택되지 않았습니다.'
        }), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({
            'success': False,
            'message': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다.'
        }), 400

    try:
        # Excel 파일 읽기
        df = pd.read_excel(file)

        # 필수 컬럼 확인
        required_columns = ['국문사업명', '국가별']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'필수 컬럼이 없습니다: {", ".join(missing_columns)}'
            }), 400

        # 컬럼 매핑
        column_mapping = {
            '번호': 'number',
            '수주년도': 'contract_year',
            '진행여부': 'status',
            '국가별': 'country',
            'X': 'longitude',
            'Y': 'latitude',
            '영문사업명': 'title_en',
            '국문사업명': 'title_kr',
            '사업형태': 'project_type',
            '착수일': 'start_date',
            '준공일': 'end_date',
            '용역비(공사)(백만원)': 'budget',
            '발주처': 'client'
        }

        imported_count = 0
        skipped_count = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                # 필수 필드 검증
                if pd.isna(row.get('국문사업명')) or pd.isna(row.get('국가별')):
                    errors.append(f'행 {idx + 2}: 필수 필드(국문사업명, 국가) 누락')
                    skipped_count += 1
                    continue

                # 프로젝트 데이터 준비
                project_data = {
                    'number': int(row['번호']) if pd.notna(row.get('번호')) else None,
                    'contract_year': int(row['수주년도']) if pd.notna(row.get('수주년도')) else None,
                    'status': row['진행여부'] if pd.notna(row.get('진행여부')) else '준공',
                    'country': str(row['국가별']).strip(),
                    'longitude': float(row['X']) if pd.notna(row.get('X')) else None,
                    'latitude': float(row['Y']) if pd.notna(row.get('Y')) else None,
                    'title_en': str(row['영문사업명']).strip() if pd.notna(row.get('영문사업명')) else None,
                    'title_kr': str(row['국문사업명']).strip(),
                    'project_type': str(row['사업형태']).strip() if pd.notna(row.get('사업형태')) else None,
                    'start_date': str(row['착수일']) if pd.notna(row.get('착수일')) else None,
                    'end_date': str(row['준공일']) if pd.notna(row.get('준공일')) else None,
                    'budget': float(row['용역비(공사)(백만원)']) if pd.notna(row.get('용역비(공사)(백만원)')) else None,
                    'client': str(row['발주처']).strip() if pd.notna(row.get('발주처')) else None,
                    'created_by': current_user.id
                }

                # 중복 체크
                existing = ConsultingProject.query.filter_by(
                    title_kr=project_data['title_kr'],
                    country=project_data['country']
                )
                if project_data['contract_year']:
                    existing = existing.filter_by(contract_year=project_data['contract_year'])

                if existing.first():
                    errors.append(f'행 {idx + 2}: 중복된 프로젝트 - {project_data["title_kr"]}')
                    skipped_count += 1
                    continue

                # 프로젝트 생성
                project = ConsultingProject(**project_data)
                db.session.add(project)
                imported_count += 1

            except Exception as e:
                errors.append(f'행 {idx + 2}: {str(e)}')
                skipped_count += 1
                continue

        # 데이터베이스에 커밋
        db.session.commit()

        # 활동 로그
        log = ActivityLog(
            user_id=current_user.id,
            action='import',
            entity_type='consulting_project',
            description=f'{current_user.name}님이 Excel 파일로 {imported_count}개의 해외기술용역 프로젝트를 업로드했습니다.',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'업로드가 완료되었습니다. (성공: {imported_count}개, 실패: {skipped_count}개)',
            'data': {
                'imported': imported_count,
                'skipped': skipped_count,
                'total': imported_count + skipped_count,
                'errors': errors[:10]  # 최대 10개의 에러만 반환
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'업로드 중 오류가 발생했습니다: {str(e)}'
        }), 500


@consulting_bp.route('/bulk-delete', methods=['POST'])
@token_required
def bulk_delete_consulting_projects(current_user):
    """여러 프로젝트를 일괄 삭제"""
    data = request.get_json()

    if not data or 'ids' not in data:
        return jsonify({
            'success': False,
            'message': '삭제할 프로젝트 ID 목록이 없습니다.'
        }), 400

    ids = data['ids']

    if not isinstance(ids, list) or len(ids) == 0:
        return jsonify({
            'success': False,
            'message': '유효한 ID 목록이 아닙니다.'
        }), 400

    try:
        # 프로젝트 조회
        projects = ConsultingProject.query.filter(ConsultingProject.id.in_(ids)).all()

        if not projects:
            return jsonify({
                'success': False,
                'message': '삭제할 프로젝트를 찾을 수 없습니다.'
            }), 404

        deleted_count = len(projects)
        project_titles = [p.title_kr for p in projects[:5]]  # 최대 5개만 기록

        # 프로젝트 삭제
        for project in projects:
            db.session.delete(project)

        # 활동 로그
        log = ActivityLog(
            user_id=current_user.id,
            action='bulk_delete',
            entity_type='consulting_project',
            description=f'{current_user.name}님이 {deleted_count}개의 해외기술용역 프로젝트를 일괄 삭제했습니다.',
            ip_address=request.remote_addr
        )
        db.session.add(log)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{deleted_count}개의 프로젝트가 삭제되었습니다.',
            'data': {
                'deleted': deleted_count,
                'titles': project_titles
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'삭제 중 오류가 발생했습니다: {str(e)}'
        }), 500


def parse_year_from_date(date_str):
    """날짜 문자열에서 연도 추출"""
    if not date_str:
        return None
    # 따옴표, 공백 제거
    clean_str = str(date_str).replace("'", "").replace("'", "").replace("'", "").strip()
    # 구분자로 분리
    parts = re.split(r'[.\-~/]', clean_str)
    if parts and parts[0]:
        try:
            yy = int(parts[0])
            # 2자리 연도 처리
            if 0 <= yy <= 99:
                return 1900 + yy if yy >= 50 else 2000 + yy
            # 4자리 연도
            if 1900 <= yy <= 2100:
                return yy
        except ValueError:
            pass
    return None


def match_project_with_profitability(project_title, profitability_records):
    """프로젝트 제목과 수익성 데이터 매칭 (부분 문자열 매칭)"""
    if not project_title:
        return []

    # 매칭을 위해 프로젝트 제목 정규화
    normalized_title = project_title.replace(' ', '').lower()

    matching_records = []
    for record in profitability_records:
        # 수익성 데이터 프로젝트명 정규화
        record_name = (record.project_name or '').replace(' ', '').lower()
        record_name_cleaned = record_name.split('-')[-1] if '-' in record_name else record_name

        # 부분 문자열 매칭 (양방향)
        if (normalized_title in record_name_cleaned or
            record_name_cleaned in normalized_title or
            any(part in normalized_title for part in record_name_cleaned.split() if len(part) > 3)):
            matching_records.append(record)

    return matching_records


# 프로젝트 ID → WBS 코드 매핑 (수동 확정) - 복수 WBS 코드 지원
PROJECT_WBS_MAPPING = {
    118: ['39101-001-04-0007'],  # 할라우강 다목적사업 → 필리핀할라우다목적
    150: ['39101-001-04-0028'],  # 은카타베이/은코타코타지구 관개사업 → 은카타베이관개사업
    154: ['39101-001-04-0020'],  # 반테민체이 관개개발 및 홍수조절사업 → 캄보디아반테민체이
    159: ['39101-001-01-0084', '39101-001-04-0033'],  # 쉬레밸리 관개사업 설계 및 시공감리 (2020: 0084, 2021~: 0033)
    160: ['39101-001-04-0036'],  # 지하수 관개 및 농촌개발 사업 → 에티오피아지하수
    161: ['39101-001-04-0037'],  # 관개 현대화 사업 SIMURP COMPONENT C → 인니관개현대화C
    162: ['39101-001-04-0039'],  # 관개 현대화 사업 SIMURP COMPONENT A → 인니관개현대화A
    163: ['39101-001-04-0040'],  # 마뗑겡 다목적댐 건설사업 → 인니마뗑겡다목적댐사업
    164: ['39101-001-04-0041'],  # 수자원관리 정보화사업 → 네팔수자원관리정보화사업
    165: ['39101-001-04-0043'],  # 까리안 다목적댐 건설사업 2단계 → 인니까리안댐(2단계)
    166: ['39101-001-04-0042'],  # 잔지바르 관개시설 건설사업 2단계 → 잔지바르관개시설(2단계)
    167: ['39101-001-04-0045'],  # 중규모관개사업 → 말라위중규모관개사업
    169: ['39101-001-04-0046'],  # 바벨지구 관개개발 및 홍수피해저감사업 → 캄보디아바벨지구관개개발
    170: ['39101-001-04-0047'],  # 지속가능한 어업을 위한 수산커뮤니티 개발 → 키리바시수산개발
    # 174: 라오스 사업은 WBS 없음 (신규사업)
}


@consulting_bp.route('/current-year-projects', methods=['GET'])
@token_required
def get_current_year_projects(current_user):
    """금년도 추진사업 목록 - 연도별 수익 데이터 포함 (WBS 매핑 기반)"""
    try:
        current_year = datetime.now().year
        next_year = current_year + 1

        # 1. 시행중인 프로젝트 가져오기
        all_projects = ConsultingProject.query.all()

        in_progress_projects = []
        for project in all_projects:
            _, effective_status = calculate_progress(project)
            if effective_status == '시행중':
                in_progress_projects.append(project)

        if not in_progress_projects:
            return jsonify({
                'success': True,
                'currentYear': current_year,
                'nextYear': next_year,
                'yearColumns': [],
                'projects': [],
                'totals': {}
            })

        # 2. 연도 범위 계산 및 프로젝트별 수익 데이터 구성
        all_years = set()
        project_data_list = []

        for project in in_progress_projects:
            start_year = parse_year_from_date(project.start_date)
            end_year = parse_year_from_date(project.end_date)

            if not start_year:
                start_year = project.contract_year or current_year
            if not end_year:
                end_year = current_year

            # 시작연도부터 금년도까지의 연도 추가
            for y in range(start_year, current_year + 1):
                all_years.add(y)

            # WBS 매핑을 통해 수익성 데이터 가져오기 (복수 WBS 코드 지원)
            wbs_codes = PROJECT_WBS_MAPPING.get(project.id, [])
            revenues_by_year = {}
            profits_by_year = {}  # 연도별 손익 추가
            costs_by_year = {}  # 연도별 비용 (직접비+인건비+경비)
            total_past_revenue = 0
            total_past_profit = 0  # 총 손익 합계
            total_past_cost = 0  # 총 비용 합계

            if wbs_codes:
                # 모든 WBS 코드에 대해 수익성 데이터 조회
                profitability_records = ProfitabilityData.query.filter(
                    ProfitabilityData.wbs_code.in_(wbs_codes)
                ).all()

                for record in profitability_records:
                    year = record.year
                    revenue = int(record.revenue or 0)
                    profit = int(record.profit or 0)  # 손익 데이터
                    # 비용 = 직접비 + 인건비 + 경비
                    direct_cost = int(record.direct_cost or 0)
                    labor_cost = int(record.labor_cost or 0)
                    expense = int(record.expense or 0)
                    total_cost = direct_cost + labor_cost + expense

                    if year not in revenues_by_year:
                        revenues_by_year[year] = 0
                    revenues_by_year[year] += revenue
                    total_past_revenue += revenue

                    if year not in profits_by_year:
                        profits_by_year[year] = 0
                    profits_by_year[year] += profit
                    total_past_profit += profit

                    if year not in costs_by_year:
                        costs_by_year[year] = 0
                    costs_by_year[year] += total_cost
                    total_past_cost += total_cost

            # 내년도 이후 수익 계산: 공사지분 용역비 - 그간 수익의 합
            # 용역비는 백만원 단위, 수익은 천원 단위이므로 변환 필요
            budget = float(project.budget) if project.budget else 0
            budget_in_thousands = budget * 1000  # 백만원 → 천원 변환

            krc_budget = float(project.krc_budget) if project.krc_budget else 0
            krc_budget_in_thousands = krc_budget * 1000  # 백만원 → 천원 변환
            krc_share_ratio = float(project.krc_share_ratio) if project.krc_share_ratio else 0
            future_revenue = max(0, krc_budget_in_thousands - total_past_revenue)

            project_data_list.append({
                'id': project.id,
                'titleKr': project.title_kr,
                'country': project.country,
                'startYear': start_year,
                'endYear': end_year,
                'budget': budget,  # 전체용역비 (백만원 단위)
                'budgetInThousands': budget_in_thousands,  # 전체용역비 (천원 단위)
                'krcBudget': krc_budget,  # 공사지분 용역비 (백만원 단위)
                'krcBudgetInThousands': krc_budget_in_thousands,  # 공사지분 용역비 (천원 단위)
                'krcShareRatio': krc_share_ratio,  # 공사지분율 (0.204 = 20.4%)
                'revenues': revenues_by_year,  # 천원 단위
                'profits': profits_by_year,  # 연도별 손익 (천원 단위)
                'costs': costs_by_year,  # 연도별 비용 (직접비+인건비+경비) (천원 단위)
                'totalPastRevenue': total_past_revenue,  # 천원 단위
                'totalPastProfit': total_past_profit,  # 총 손익 합계 (천원 단위)
                'totalPastCost': total_past_cost,  # 총 비용 합계 (천원 단위)
                'futureRevenue': future_revenue,  # 천원 단위
                'wbsCodes': wbs_codes  # 복수 WBS 코드 리스트
            })

        # 3. 연도 정렬 (오름차순)
        year_columns = sorted(list(all_years))

        # 4. 연도별 합계 계산 (수익 + 손익 + 비용)
        totals_by_year = {y: 0 for y in year_columns}
        profits_by_year_totals = {y: 0 for y in year_columns}  # 연도별 손익 합계
        costs_by_year_totals = {y: 0 for y in year_columns}  # 연도별 비용 합계
        total_future = 0
        total_budget_sum = 0  # 전체용역비 합계
        total_krc_budget_sum = 0  # 공사지분 합계
        total_past_revenue_sum = 0
        total_past_profit_sum = 0  # 총 손익 합계
        total_past_cost_sum = 0  # 총 비용 합계

        for project_data in project_data_list:
            for year, revenue in project_data['revenues'].items():
                if year in totals_by_year:
                    totals_by_year[year] += revenue
            for year, profit in project_data['profits'].items():
                if year in profits_by_year_totals:
                    profits_by_year_totals[year] += profit
            for year, cost in project_data['costs'].items():
                if year in costs_by_year_totals:
                    costs_by_year_totals[year] += cost
            total_future += project_data['futureRevenue']
            total_budget_sum += project_data['budgetInThousands']
            total_krc_budget_sum += project_data['krcBudgetInThousands']
            total_past_revenue_sum += project_data['totalPastRevenue']
            total_past_profit_sum += project_data['totalPastProfit']
            total_past_cost_sum += project_data['totalPastCost']

        return jsonify({
            'success': True,
            'currentYear': current_year,
            'nextYear': next_year,
            'yearColumns': year_columns,
            'projects': project_data_list,
            'totals': {
                'byYear': totals_by_year,
                'profitsByYear': profits_by_year_totals,  # 연도별 손익 합계
                'costsByYear': costs_by_year_totals,  # 연도별 비용 합계 (직접비+인건비+경비)
                'future': total_future,
                'totalBudget': total_budget_sum,  # 전체용역비 합계 (천원 단위)
                'totalKrcBudget': total_krc_budget_sum,  # 공사지분 총 용역비 (천원 단위)
                'totalPastRevenue': total_past_revenue_sum,
                'totalPastProfit': total_past_profit_sum,  # 총 손익 합계
                'totalPastCost': total_past_cost_sum  # 총 비용 합계
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'금년도 추진사업 목록 조회 중 오류가 발생했습니다: {str(e)}'
        }), 500


@consulting_bp.route('/<int:consulting_id>/personnel', methods=['POST'])
@token_required
def add_consulting_personnel(current_user, consulting_id):
    """컨설팅 프로젝트 인력 추가"""
    try:
        project = ConsultingProject.query.get_or_404(consulting_id)
        
        data = request.get_json()
        
        # 필수 필드 검증
        if not data.get('name'):
            return jsonify({'success': False, 'message': '이름은 필수 항목입니다.'}), 400
        
        # 날짜 형식 변환
        start_date = None
        end_date = None
        
        if data.get('startDate'):
            try:
                s = data['startDate'].replace('.', '-')
                start_date = datetime.strptime(s[:10], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'message': '시작일 형식이 올바르지 않습니다. (YYYY.MM.DD)'}), 400

        if data.get('endDate'):
            try:
                s = data['endDate'].replace('.', '-')
                end_date = datetime.strptime(s[:10], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'message': '종료일 형식이 올바르지 않습니다. (YYYY.MM.DD)'}), 400
        
        # 인력 생성
        personnel = ConsultingPersonnel(
            consulting_project_id=consulting_id,
            name=data['name'],
            role=data.get('role'),
            position=data.get('position'),
            affiliation=data.get('affiliation'),
            start_date=start_date,
            end_date=end_date,
            is_deployed=data.get('isDeployed', False),
            contact_email=data.get('contactEmail'),
            contact_phone=data.get('contactPhone')
        )
        
        db.session.add(personnel)
        
        # 활동 로그
        log = ActivityLog(
            user_id=current_user.id,
            action='create',
            entity_type='consulting_personnel',
            entity_id=None,
            description=f'컨설팅 프로젝트 인력 추가: {project.title_kr} - {data["name"]}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': '인력이 추가되었습니다.',
            'data': personnel.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'인력 추가 실패: {str(e)}'}), 500


@consulting_bp.route('/personnel/<int:personnel_id>', methods=['PUT'])
@token_required
def update_consulting_personnel(current_user, personnel_id):
    """컨설팅 프로젝트 인력 수정"""
    personnel = ConsultingPersonnel.query.get_or_404(personnel_id)
    data = request.get_json()

    if 'name' in data:
        personnel.name = data['name']
    if 'role' in data:
        personnel.role = data['role']
    if 'startDate' in data:
        personnel.start_date = datetime.strptime(data['startDate'].replace('.', '-')[:10], '%Y-%m-%d').date() if data['startDate'] else None
    if 'endDate' in data:
        personnel.end_date = datetime.strptime(data['endDate'].replace('.', '-')[:10], '%Y-%m-%d').date() if data['endDate'] else None
    if 'isDeployed' in data:
        personnel.is_deployed = data['isDeployed']

    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action='update',
        entity_type='consulting_personnel',
        entity_id=personnel.id,
        description=f'컨설팅 파견정보 수정: {personnel.name}',
        ip_address=request.remote_addr
    )
    db.session.add(log)

    db.session.commit()

    return jsonify({
        'success': True,
        'message': '파견정보가 수정되었습니다.',
        'data': personnel.to_dict()
    })


@consulting_bp.route('/personnel/<int:personnel_id>', methods=['DELETE'])
@token_required
def delete_consulting_personnel(current_user, personnel_id):
    """컨설팅 프로젝트 인력 삭제"""
    personnel = ConsultingPersonnel.query.get_or_404(personnel_id)
    personnel_name = personnel.name

    # Log activity
    log = ActivityLog(
        user_id=current_user.id,
        action='delete',
        entity_type='consulting_personnel',
        entity_id=personnel.id,
        description=f'컨설팅 파견정보 삭제: {personnel_name}',
        ip_address=request.remote_addr
    )
    db.session.add(log)

    db.session.delete(personnel)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': '파견정보가 삭제되었습니다.'
    })


# ==========================================
# 사업제안 및 수주현황 API
# ==========================================

@consulting_bp.route('/proposal-status', methods=['GET'])
@token_required
def get_proposal_statuses(current_user):
    """사업제안 및 수주현황 목록 조회"""
    items = ProposalStatus.query.order_by(ProposalStatus.sort_order, ProposalStatus.id).all()
    return jsonify({
        'success': True,
        'data': {'items': [item.to_dict() for item in items]}
    })


@consulting_bp.route('/proposal-status', methods=['POST'])
@token_required
def create_proposal_status(current_user):
    """사업제안 수주현황 추가"""
    if current_user.role not in ('admin', 'manager'):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    data = request.get_json()
    item = ProposalStatus(
        project_name=data.get('projectName', ''),
        funding=data.get('funding', ''),
        sort_order=data.get('sortOrder', 0),
        consortium=data.get('consortium', ''),
        share_ratio=data.get('shareRatio', ''),
        eoi_date=normalize_date_dot(data.get('eoiDate', '')),
        shortlist_date=normalize_date_dot(data.get('shortlistDate', '')),
        announcement_date=normalize_date_dot(data.get('announcementDate', '')),
        proposal_date=normalize_date_dot(data.get('proposalDate', '')),
        selection_date=normalize_date_dot(data.get('selectionDate', '')),
        negotiation_date=normalize_date_dot(data.get('negotiationDate', '')),
        contract_date=normalize_date_dot(data.get('contractDate', '')),
        eoi_progress=data.get('eoiProgress', False),
        shortlist_progress=data.get('shortlistProgress', False),
        announcement_progress=data.get('announcementProgress', False),
        proposal_progress=data.get('proposalProgress', False),
        selection_progress=data.get('selectionProgress', False),
        negotiation_progress=data.get('negotiationProgress', False),
        contract_progress=data.get('contractProgress', False),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify({'success': True, 'data': item.to_dict(), 'message': '추가되었습니다.'})


@consulting_bp.route('/proposal-status/<int:id>', methods=['PUT'])
@token_required
def update_proposal_status(current_user, id):
    """사업제안 수주현황 수정"""
    if current_user.role not in ('admin', 'manager'):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    item = ProposalStatus.query.get_or_404(id)
    data = request.get_json()

    item.project_name = data.get('projectName', item.project_name)
    item.funding = data.get('funding', item.funding)
    item.sort_order = data.get('sortOrder', item.sort_order)
    item.consortium = data.get('consortium', item.consortium)
    item.share_ratio = data.get('shareRatio', item.share_ratio)
    item.eoi_date = data.get('eoiDate', item.eoi_date)
    item.shortlist_date = data.get('shortlistDate', item.shortlist_date)
    item.announcement_date = data.get('announcementDate', item.announcement_date)
    item.proposal_date = data.get('proposalDate', item.proposal_date)
    item.selection_date = data.get('selectionDate', item.selection_date)
    item.negotiation_date = data.get('negotiationDate', item.negotiation_date)
    item.contract_date = data.get('contractDate', item.contract_date)
    item.eoi_progress = data.get('eoiProgress', item.eoi_progress)
    item.shortlist_progress = data.get('shortlistProgress', item.shortlist_progress)
    item.announcement_progress = data.get('announcementProgress', item.announcement_progress)
    item.proposal_progress = data.get('proposalProgress', item.proposal_progress)
    item.selection_progress = data.get('selectionProgress', item.selection_progress)
    item.negotiation_progress = data.get('negotiationProgress', item.negotiation_progress)
    item.contract_progress = data.get('contractProgress', item.contract_progress)

    db.session.commit()
    return jsonify({'success': True, 'data': item.to_dict(), 'message': '수정되었습니다.'})


@consulting_bp.route('/proposal-status/bulk', methods=['PUT'])
@token_required
def bulk_update_proposal_status(current_user):
    """사업제안 수주현황 일괄 수정"""
    if current_user.role not in ('admin', 'manager'):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    data = request.get_json()
    items = data.get('items', [])

    for item_data in items:
        item_id = item_data.get('id')
        if item_id:
            item = ProposalStatus.query.get(item_id)
            if item:
                item.project_name = item_data.get('projectName', item.project_name)
                item.funding = item_data.get('funding', item.funding)
                item.sort_order = item_data.get('sortOrder', item.sort_order)
                item.consortium = item_data.get('consortium', item.consortium)
                item.share_ratio = item_data.get('shareRatio', item.share_ratio)
                item.eoi_date = normalize_date_dot(item_data.get('eoiDate', item.eoi_date))
                item.shortlist_date = normalize_date_dot(item_data.get('shortlistDate', item.shortlist_date))
                item.announcement_date = normalize_date_dot(item_data.get('announcementDate', item.announcement_date))
                item.proposal_date = normalize_date_dot(item_data.get('proposalDate', item.proposal_date))
                item.selection_date = normalize_date_dot(item_data.get('selectionDate', item.selection_date))
                item.negotiation_date = normalize_date_dot(item_data.get('negotiationDate', item.negotiation_date))
                item.contract_date = normalize_date_dot(item_data.get('contractDate', item.contract_date))
                item.eoi_progress = item_data.get('eoiProgress', item.eoi_progress)
                item.shortlist_progress = item_data.get('shortlistProgress', item.shortlist_progress)
                item.announcement_progress = item_data.get('announcementProgress', item.announcement_progress)
                item.proposal_progress = item_data.get('proposalProgress', item.proposal_progress)
                item.selection_progress = item_data.get('selectionProgress', item.selection_progress)
                item.negotiation_progress = item_data.get('negotiationProgress', item.negotiation_progress)
                item.contract_progress = item_data.get('contractProgress', item.contract_progress)
        else:
            # 새 항목 추가
            new_item = ProposalStatus(
                project_name=item_data.get('projectName', ''),
                funding=item_data.get('funding', ''),
                sort_order=item_data.get('sortOrder', 0),
                consortium=item_data.get('consortium', ''),
                share_ratio=item_data.get('shareRatio', ''),
                eoi_date=normalize_date_dot(item_data.get('eoiDate', '')),
                shortlist_date=normalize_date_dot(item_data.get('shortlistDate', '')),
                announcement_date=normalize_date_dot(item_data.get('announcementDate', '')),
                proposal_date=normalize_date_dot(item_data.get('proposalDate', '')),
                selection_date=normalize_date_dot(item_data.get('selectionDate', '')),
                negotiation_date=normalize_date_dot(item_data.get('negotiationDate', '')),
                contract_date=normalize_date_dot(item_data.get('contractDate', '')),
                eoi_progress=item_data.get('eoiProgress', False),
                shortlist_progress=item_data.get('shortlistProgress', False),
                announcement_progress=item_data.get('announcementProgress', False),
                proposal_progress=item_data.get('proposalProgress', False),
                selection_progress=item_data.get('selectionProgress', False),
                negotiation_progress=item_data.get('negotiationProgress', False),
                contract_progress=item_data.get('contractProgress', False),
            )
            db.session.add(new_item)

    db.session.commit()
    items = ProposalStatus.query.order_by(ProposalStatus.sort_order, ProposalStatus.id).all()
    return jsonify({'success': True, 'data': {'items': [i.to_dict() for i in items]}, 'message': '저장되었습니다.'})


@consulting_bp.route('/proposal-status/<int:id>', methods=['DELETE'])
@token_required
def delete_proposal_status(current_user, id):
    """사업제안 수주현황 삭제"""
    if current_user.role not in ('admin', 'manager'):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    item = ProposalStatus.query.get_or_404(id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'success': True, 'message': '삭제되었습니다.'})


# ============================================================
# 사업 라이프사이클 API
# ============================================================

def get_lifecycle_stages_from_project(project):
    """프로젝트의 체크박스 필드 기반으로 라이프사이클 단계 구성 반환"""
    stages = [
        {'key': 'eoi',        'label': 'EOI제출'},
        {'key': 'proposal',   'label': '제안서제출'},
        {'key': 'contract',   'label': '계약'},
        {'key': 'completion', 'label': '준공'},
    ]
    return stages


def format_date_short(date_val):
    """날짜를 'YY.MM 형식으로 변환"""
    if not date_val:
        return ''
    s = str(date_val)
    try:
        # '2025-03-06' 형태
        if '-' in s and len(s) >= 7:
            parts = s.split('-')
            return f"'{parts[0][2:]}.{parts[1]}"
        # '2026-02-09 04:36:22' datetime 형태
        if ' ' in s:
            s = s.split(' ')[0]
            parts = s.split('-')
            return f"'{parts[0][2:]}.{parts[1]}"
    except Exception:
        pass
    return s[:8] if len(s) > 8 else s


def build_lifecycle_data(project, lifecycle_record, eoi_list=None, proposal_list=None, contract_list=None, perf_list=None):
    """프로젝트와 라이프사이클 레코드 + 관련 모델 데이터에서 단계별 상태를 구성"""
    stage_defs = get_lifecycle_stages_from_project(project)
    lc = lifecycle_record

    # 1단계: lifecycle_record에서 수동 입력 데이터 로드
    stage_data_map = {}
    if lc:
        stage_data_map = {
            'eoi':          {'date': lc.eoi_date,          'completed': lc.eoi_completed,          'progress': lc.eoi_progress},
            'shortlist':    {'date': lc.shortlist_date,    'completed': lc.shortlist_completed,    'progress': lc.shortlist_progress},
            'proposal':     {'date': lc.proposal_date,     'completed': lc.proposal_completed,     'progress': lc.proposal_progress},
            'contract':     {'date': lc.contract_date,     'completed': lc.contract_completed,     'progress': lc.contract_progress},
            'kickoff':      {'date': lc.kickoff_date,      'completed': lc.kickoff_completed,      'progress': lc.kickoff_progress},
            'design':       {'date': lc.design_date,       'completed': lc.design_completed,       'progress': lc.design_progress},
            'construction': {'date': lc.construction_date, 'completed': lc.construction_completed, 'progress': lc.construction_progress},
            'completion':   {'date': lc.completion_date,   'completed': lc.completion_completed,   'progress': lc.completion_progress},
        }

    # 2단계: 관련 모델 데이터로 보강 (lifecycle_record에 값이 없을 때만)
    # 파일 정보 사전 추출
    eoi_file_info = {'hasFile': False, 'fileId': None, 'fileName': None}
    prop_file_info = {'hasFile': False, 'fileId': None, 'fileName': None}
    cont_file_info = {'hasFile': False, 'fileId': None, 'fileName': None}
    compl_file_info = {'hasFile': False, 'fileId': None, 'fileName': None}

    if eoi_list:
        latest_eoi = eoi_list[-1]
        if latest_eoi and latest_eoi.eoi_file_path:
            eoi_display = f"EOI 제출서 ({format_date_short(latest_eoi.submission_date)})" if latest_eoi.submission_date else 'EOI 제출서'
            eoi_file_info = {'hasFile': True, 'fileId': latest_eoi.id, 'fileName': latest_eoi.eoi_file_name or 'EOI 제출서', 'fileDisplayName': eoi_display}

    if proposal_list:
        latest_prop = proposal_list[-1]
        has_any_file = latest_prop and (latest_prop.file_path or latest_prop.technical_file_path or latest_prop.price_file_path)
        if has_any_file:
            prop_display = f"기술제안서 ({format_date_short(latest_prop.submission_date)})" if latest_prop.submission_date else '기술제안서'
            file_name = latest_prop.technical_file_name or latest_prop.file_name or '기술제안서'
            prop_file_info = {'hasFile': True, 'fileId': latest_prop.id, 'fileName': file_name, 'fileDisplayName': prop_display}

    if contract_list:
        contracts_with_file = [c for c in contract_list if c.document_type == 'contract' and c.file_path]
        if contracts_with_file:
            latest_cont = contracts_with_file[-1]
            cont_order = len(contracts_with_file)
            cont_display = f"{cont_order}차 계약서" if cont_order > 0 else '계약서'
            cont_file_info = {'hasFile': True, 'fileId': latest_cont.id, 'fileName': latest_cont.file_name or '계약서', 'fileDisplayName': cont_display}

        # 준공보고서 (final_report)
        final_reports = [c for c in contract_list if c.document_type == 'final_report' and c.file_path]
        if final_reports:
            latest_final = final_reports[-1]
            compl_file_info = {'hasFile': True, 'fileId': latest_final.id, 'fileName': latest_final.file_name or '최종보고서', 'fileDisplayName': '최종보고서'}

    file_map = {
        'eoi': eoi_file_info,
        'proposal': prop_file_info,
        'contract': cont_file_info,
        'completion': compl_file_info,
    }

    # --- EOI ---
    eoi_data = stage_data_map.get('eoi', {})
    if not eoi_data.get('date'):
        if eoi_list:
            # 가장 최신 EOI
            latest_eoi = eoi_list[-1] if eoi_list else None
            if latest_eoi:
                eoi_date = format_date_short(latest_eoi.submission_date)
                eoi_result = (latest_eoi.result or '').strip()
                is_completed = eoi_result in ('통과', '선정')
                is_progress = eoi_result in ('대기중', '심사중', '')
                stage_data_map['eoi'] = {
                    'date': eoi_date,
                    'completed': is_completed or bool(eoi_date and not is_progress),
                    'progress': is_progress and bool(eoi_date)
                }
    # --- 제안서 ---
    prop_data = stage_data_map.get('proposal', {})
    if not prop_data.get('date'):
        if proposal_list:
            latest_prop = proposal_list[-1] if proposal_list else None
            if latest_prop:
                prop_date = format_date_short(latest_prop.submission_date)
                prop_result = (latest_prop.result or '').strip()
                is_completed = prop_result == '선정'
                is_progress = prop_result in ('심사중', '')
                stage_data_map['proposal'] = {
                    'date': prop_date,
                    'completed': is_completed or bool(prop_date and not is_progress),
                    'progress': is_progress and bool(prop_date)
                }

    # --- 계약 ---
    cont_data = stage_data_map.get('contract', {})
    if not cont_data.get('date'):
        if contract_list:
            # contract 타입인 것 중 가장 최신
            contracts = [c for c in contract_list if c.document_type == 'contract']
            if contracts:
                latest_contract = contracts[-1]
                cont_date = format_date_short(latest_contract.upload_date)
                stage_data_map['contract'] = {'date': cont_date, 'completed': True, 'progress': False}

    # --- 착수 (ConsultingProject.start_date) ---
    kickoff_data = stage_data_map.get('kickoff', {})
    if not kickoff_data.get('date'):
        if project.start_date:
            kickoff_date = format_date_short(project.start_date)
            stage_data_map['kickoff'] = {'date': kickoff_date, 'completed': True, 'progress': False}

    # --- 준공 (ConsultingProject.end_date 또는 PerformanceRecord.end_date) ---
    comp_data = stage_data_map.get('completion', {})
    if not comp_data.get('date'):
        comp_date = ''
        # PerformanceRecord에서 end_date 확인
        if perf_list:
            for pr in perf_list:
                if pr.end_date:
                    comp_date = format_date_short(pr.end_date)
                    break
        # 없으면 ConsultingProject.end_date
        if not comp_date and project.end_date:
            comp_date = format_date_short(project.end_date)
        if comp_date and project.status == '준공':
            stage_data_map['completion'] = {'date': comp_date, 'completed': True, 'progress': False}

    # 3단계: 최종 단계 리스트 생성
    stages = []
    for sd in stage_defs:
        key = sd['key']
        data = stage_data_map.get(key, {})
        date = data.get('date', '') or ''
        progress = data.get('progress', False)
        completed = data.get('completed', False)

        if progress:
            status = 'in-progress'
        elif completed or date:
            status = 'completed'
        else:
            status = 'pending'

        file_info = file_map.get(key, {'hasFile': False, 'fileId': None, 'fileName': None})
        stages.append({
            'key': key,
            'label': sd['label'],
            'date': date,
            'status': status,
            'hasFile': file_info['hasFile'],
            'fileId': file_info['fileId'],
            'fileName': file_info['fileName'],
            'fileDisplayName': file_info.get('fileDisplayName'),
        })

    # 자동 완성: 후속 단계가 completed/in-progress면 이전 pending 단계를 completed-nodata로
    for i in range(len(stages) - 1, -1, -1):
        if stages[i]['status'] in ('completed', 'in-progress'):
            for j in range(i):
                if stages[j]['status'] == 'pending':
                    stages[j]['status'] = 'completed-nodata'
            break

    # 준공 상태인 프로젝트: 날짜 없는 단계는 completed-nodata로
    if project.status == '준공':
        for s in stages:
            if s['status'] == 'pending':
                s['status'] = 'completed-nodata' if not s['date'] else 'completed'

    # 날짜가 있는 completed는 그대로, 날짜 없는 completed는 completed-nodata로 변환
    for s in stages:
        if s['status'] == 'completed' and not s['date']:
            s['status'] = 'completed-nodata'

    # 날짜 역순 감지: 이전 단계보다 날짜가 앞서면 dateError 플래그
    def parse_date_for_compare(date_str):
        """'YY.MM 형식을 비교 가능한 숫자로 변환"""
        if not date_str:
            return None
        d = date_str.strip().lstrip("'")
        parts = d.split('.')
        if len(parts) >= 2:
            try:
                y = int(parts[0])
                m = int(parts[1])
                # 2자리 연도: 70 이상이면 1900년대, 미만이면 2000년대
                full_y = (1900 + y) if y >= 70 else (2000 + y)
                return full_y * 100 + m
            except ValueError:
                pass
        return None

    prev_date_val = None
    for s in stages:
        cur_val = parse_date_for_compare(s['date'])
        if cur_val is not None and prev_date_val is not None:
            if cur_val < prev_date_val:
                s['dateError'] = True
        if cur_val is not None:
            prev_date_val = cur_val

    return stages


def resolve_file_path(stored_path, subfolder):
    """크로스 플랫폼 파일 경로 해석.
    DB에 Windows 절대 경로가 저장되어 있어도 로컬 uploads/ 하위에서 파일을 찾는다."""
    if not stored_path:
        return None
    if os.path.exists(stored_path):
        return stored_path
    filename = stored_path.replace('\\', '/').split('/')[-1]
    upload_folder = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        'uploads', subfolder
    )
    local_path = os.path.join(upload_folder, filename)
    if os.path.exists(local_path):
        return local_path
    return None


@consulting_bp.route('/<int:project_id>/lifecycle/<stage>/file', methods=['GET'])
@token_required
def download_lifecycle_file(current_user, project_id, stage):
    """라이프사이클 단계별 첨부 파일 다운로드"""
    ConsultingProject.query.get_or_404(project_id)

    file_path = None
    file_name = None

    if stage == 'eoi':
        eoi = Eoi.query.filter_by(consulting_project_id=project_id).order_by(Eoi.id.desc()).first()
        if eoi and eoi.eoi_file_path:
            file_path = eoi.eoi_file_path
            file_name = eoi.eoi_file_name
    elif stage == 'proposal':
        prop = Proposal.query.filter_by(consulting_project_id=project_id).order_by(Proposal.id.desc()).first()
        if prop and prop.file_path:
            file_path = prop.file_path
            file_name = prop.file_name
    elif stage == 'contract':
        cont = Contract.query.filter_by(
            consulting_project_id=project_id, document_type='contract'
        ).order_by(Contract.id.desc()).first()
        if cont and cont.file_path:
            file_path = cont.file_path
            file_name = cont.file_name
    elif stage == 'completion':
        final = Contract.query.filter_by(
            consulting_project_id=project_id, document_type='final_report'
        ).order_by(Contract.id.desc()).first()
        if final and final.file_path:
            file_path = final.file_path
            file_name = final.file_name

    if not file_path:
        return jsonify({'success': False, 'message': '첨부 파일이 없습니다.'}), 404

    subfolder_map = {'eoi': 'eoi', 'proposal': 'proposals', 'contract': 'contracts', 'completion': 'contracts'}
    subfolder = subfolder_map.get(stage, 'documents')
    resolved = resolve_file_path(file_path, subfolder)

    if not resolved or not os.path.exists(resolved):
        return jsonify({'success': False, 'message': '파일을 찾을 수 없습니다.'}), 404

    safe_name = file_name or os.path.basename(resolved)
    return send_file(resolved, as_attachment=True, download_name=safe_name)


@consulting_bp.route('/<int:id>/lifecycle', methods=['GET'])
@token_required
def get_project_lifecycle(current_user, id):
    """단일 프로젝트 라이프사이클 조회"""
    project = ConsultingProject.query.get_or_404(id)
    lc = ProjectLifecycle.query.filter_by(consulting_project_id=id).first()

    eoi_list = Eoi.query.filter_by(consulting_project_id=id).order_by(Eoi.id.asc()).all()
    proposal_list = Proposal.query.filter_by(consulting_project_id=id).order_by(Proposal.id.asc()).all()
    contract_list = Contract.query.filter_by(consulting_project_id=id).order_by(Contract.id.asc()).all()
    perf_list = PerformanceRecord.query.filter_by(consulting_project_id=id).order_by(PerformanceRecord.id.asc()).all()

    stages = build_lifecycle_data(project, lc,
        eoi_list=eoi_list, proposal_list=proposal_list,
        contract_list=contract_list, perf_list=perf_list)

    return jsonify({'success': True, 'data': {'stages': stages}})


@consulting_bp.route('/filter-options', methods=['GET'])
@token_required
def get_filter_options(current_user):
    """필터 드롭다운에 사용할 고유 값 목록 반환"""
    clients = db.session.query(ConsultingProject.client).filter(
        ConsultingProject.client.isnot(None), ConsultingProject.client != ''
    ).distinct().order_by(ConsultingProject.client).all()

    funding_sources = db.session.query(ConsultingProject.funding_source).filter(
        ConsultingProject.funding_source.isnot(None), ConsultingProject.funding_source != ''
    ).distinct().order_by(ConsultingProject.funding_source).all()

    return jsonify({
        'success': True,
        'data': {
            'clients': [c[0] for c in clients],
            'fundingSources': [f[0] for f in funding_sources],
        }
    })


@consulting_bp.route('/lifecycle-list', methods=['GET'])
@token_required
def get_lifecycle_list(current_user):
    """사업 라이프사이클 목록 조회"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    country = request.args.get('country', '')
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    project_type_filter = request.args.get('project_type', '')
    year = request.args.get('year', None, type=int)
    client = request.args.get('client', '')
    funding_source = request.args.get('funding_source', '')

    query = ConsultingProject.query

    if year:
        query = query.filter(ConsultingProject.contract_year == year)
    if country:
        query = query.filter(ConsultingProject.country == country)
    if status:
        query = query.filter(ConsultingProject.status == status)
    if client:
        query = query.filter(ConsultingProject.client == client)
    if funding_source:
        query = query.filter(ConsultingProject.funding_source == funding_source)
    if search:
        query = query.filter(db.or_(
            ConsultingProject.title_kr.ilike(f'%{search}%'),
            ConsultingProject.title_en.ilike(f'%{search}%')
        ))
    if project_type_filter:
        type_filter_map = {
            'feasibility': ConsultingProject.type_feasibility,
            'masterplan': ConsultingProject.type_masterplan,
            'basic_design': ConsultingProject.type_basic_design,
            'detailed_design': ConsultingProject.type_detailed_design,
            'construction': ConsultingProject.type_construction,
            'pmc': ConsultingProject.type_pmc,
        }
        if project_type_filter in type_filter_map:
            query = query.filter(type_filter_map[project_type_filter] == True)
        else:
            query = query.filter(ConsultingProject.project_type.ilike(f'%{project_type_filter}%'))

    query = query.order_by(ConsultingProject.contract_year.desc(), ConsultingProject.id.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    projects = pagination.items

    # 일괄 조회: 해당 페이지 프로젝트들의 관련 데이터
    project_ids = [p.id for p in projects]
    lifecycle_records = {}
    eoi_map = {}
    proposal_map = {}
    contract_map = {}
    perf_map = {}

    if project_ids:
        # ProjectLifecycle
        records = ProjectLifecycle.query.filter(ProjectLifecycle.consulting_project_id.in_(project_ids)).all()
        lifecycle_records = {r.consulting_project_id: r for r in records}

        # EOI (consulting_project_id 기준)
        eois = Eoi.query.filter(Eoi.consulting_project_id.in_(project_ids)).order_by(Eoi.id.asc()).all()
        for e in eois:
            eoi_map.setdefault(e.consulting_project_id, []).append(e)

        # Proposal
        proposals = Proposal.query.filter(Proposal.consulting_project_id.in_(project_ids)).order_by(Proposal.id.asc()).all()
        for pr in proposals:
            proposal_map.setdefault(pr.consulting_project_id, []).append(pr)

        # Contract
        contracts = Contract.query.filter(Contract.consulting_project_id.in_(project_ids)).order_by(Contract.id.asc()).all()
        for c in contracts:
            contract_map.setdefault(c.consulting_project_id, []).append(c)

        # PerformanceRecord
        perfs = PerformanceRecord.query.filter(PerformanceRecord.consulting_project_id.in_(project_ids)).order_by(PerformanceRecord.id.asc()).all()
        for pr in perfs:
            perf_map.setdefault(pr.consulting_project_id, []).append(pr)

    result = []
    for p in projects:
        lc = lifecycle_records.get(p.id)
        stages = build_lifecycle_data(
            p, lc,
            eoi_list=eoi_map.get(p.id),
            proposal_list=proposal_map.get(p.id),
            contract_list=contract_map.get(p.id),
            perf_list=perf_map.get(p.id)
        )
        progress, _ = calculate_progress(p)

        # 컨소시엄 구성
        consortium_parts = []
        if p.lead_company:
            ratio_str = f"({int(p.lead_company_ratio * 100)}%)" if p.lead_company_ratio else ''
            consortium_parts.append(f"{p.lead_company}{ratio_str}")
        for jv_idx in range(1, 6):
            jv_name = getattr(p, f'jv{jv_idx}', None)
            jv_ratio = getattr(p, f'jv{jv_idx}_ratio', None)
            if jv_name:
                ratio_str = f"({int(jv_ratio * 100)}%)" if jv_ratio else ''
                consortium_parts.append(f"{jv_name}{ratio_str}")

        result.append({
            'id': p.id,
            'titleKr': p.title_kr,
            'titleEn': p.title_en,
            'country': p.country,
            'projectType': p.project_type,
            'projectTypeLabel': p._get_type_label(),
            'typeFeasibility': p.type_feasibility or False,
            'typeMasterplan': p.type_masterplan or False,
            'typeBasicDesign': p.type_basic_design or False,
            'typeDetailedDesign': p.type_detailed_design or False,
            'typeConstruction': p.type_construction or False,
            'typePmc': p.type_pmc or False,
            'projectTypeEtc': p.project_type_etc,
            'status': p.status,
            'contractYear': p.contract_year,
            'startDate': p.start_date,
            'endDate': p.end_date,
            'budget': float(p.budget) if p.budget else 0,
            'krcBudget': float(p.krc_budget) if p.krc_budget else 0,
            'krcShareRatio': float(p.krc_share_ratio) if p.krc_share_ratio else 0,
            'client': p.client or '',
            'fundingSource': p.funding_source or '',
            'consortium': ', '.join(consortium_parts) if consortium_parts else '',
            'progress': progress,
            'stages': stages
        })

    return jsonify({
        'success': True,
        'data': result,
        'total': pagination.total,
        'currentPage': pagination.page,
        'pages': pagination.pages,
        'perPage': per_page
    })


@consulting_bp.route('/lifecycle/bulk', methods=['PUT'])
@token_required
def update_lifecycle_bulk(current_user):
    """사업 라이프사이클 일괄 저장"""
    if current_user.role not in ('admin', 'manager'):
        return jsonify({'success': False, 'message': '권한이 없습니다.'}), 403

    data = request.get_json()
    items = data.get('items', [])

    for item_data in items:
        project_id = item_data.get('consultingProjectId')
        if not project_id:
            continue

        lc = ProjectLifecycle.query.filter_by(consulting_project_id=project_id).first()
        if not lc:
            lc = ProjectLifecycle(consulting_project_id=project_id)
            db.session.add(lc)

        # EOI ~ 계약 단계
        lc.eoi_date = item_data.get('eoiDate', lc.eoi_date)
        lc.eoi_completed = item_data.get('eoiCompleted', lc.eoi_completed)
        lc.eoi_progress = item_data.get('eoiProgress', lc.eoi_progress)
        lc.shortlist_date = item_data.get('shortlistDate', lc.shortlist_date)
        lc.shortlist_completed = item_data.get('shortlistCompleted', lc.shortlist_completed)
        lc.shortlist_progress = item_data.get('shortlistProgress', lc.shortlist_progress)
        lc.proposal_date = item_data.get('proposalDate', lc.proposal_date)
        lc.proposal_completed = item_data.get('proposalCompleted', lc.proposal_completed)
        lc.proposal_progress = item_data.get('proposalProgress', lc.proposal_progress)
        lc.contract_date = item_data.get('contractDate', lc.contract_date)
        lc.contract_completed = item_data.get('contractCompleted', lc.contract_completed)
        lc.contract_progress = item_data.get('contractProgress', lc.contract_progress)

        # 착수 ~ 준공 단계
        lc.kickoff_date = item_data.get('kickoffDate', lc.kickoff_date)
        lc.kickoff_completed = item_data.get('kickoffCompleted', lc.kickoff_completed)
        lc.kickoff_progress = item_data.get('kickoffProgress', lc.kickoff_progress)
        lc.design_date = item_data.get('designDate', lc.design_date)
        lc.design_completed = item_data.get('designCompleted', lc.design_completed)
        lc.design_progress = item_data.get('designProgress', lc.design_progress)
        lc.construction_date = item_data.get('constructionDate', lc.construction_date)
        lc.construction_completed = item_data.get('constructionCompleted', lc.construction_completed)
        lc.construction_progress = item_data.get('constructionProgress', lc.construction_progress)
        lc.completion_date = item_data.get('completionDate', lc.completion_date)
        lc.completion_completed = item_data.get('completionCompleted', lc.completion_completed)
        lc.completion_progress = item_data.get('completionProgress', lc.completion_progress)

        # 계약일/준공일을 메인 프로젝트 데이터에 동기화 (착수일 = 계약일)
        contract_date = item_data.get('contractDate') or ''
        completion_date = item_data.get('completionDate') or ''
        if contract_date or completion_date:
            project = ConsultingProject.query.get(project_id)
            if project:
                if contract_date:
                    project.start_date = contract_date.replace('.', '-')
                if completion_date:
                    project.end_date = completion_date.replace('.', '-')

    db.session.commit()
    return jsonify({'success': True, 'message': '라이프사이클이 저장되었습니다.'})


@consulting_bp.route('/<int:id>/stage-docs', methods=['GET'])
@token_required
def get_stage_docs(current_user, id):
    """단계별 관련 문서 조회 (드롭다운 lazy load용)"""
    project = ConsultingProject.query.get_or_404(id)

    result = {
        'eoi': {},
        'proposal': {},
        'contract': {},
        'completion': {}
    }

    # EOI
    eoi = Eoi.query.filter_by(consulting_project_id=id).order_by(Eoi.id.desc()).first()
    if eoi:
        result['eoi']['eoiFile'] = {
            'eoiId': eoi.id,
            'fileName': eoi.eoi_file_name or '',
            'hasFile': bool(eoi.eoi_file_path),
            'submissionDate': eoi.submission_date.strftime("'%y.%m") if eoi.submission_date else None
        }

    # Proposal (기술제안서, 가격제안서)
    proposal = Proposal.query.filter_by(consulting_project_id=id).order_by(Proposal.id.desc()).first()
    if proposal:
        result['proposal']['technical'] = {
            'proposalId': proposal.id,
            'fileName': proposal.technical_file_name or '',
            'hasFile': bool(proposal.technical_file_path)
        }
        result['proposal']['price'] = {
            'proposalId': proposal.id,
            'fileName': proposal.price_file_name or '',
            'hasFile': bool(proposal.price_file_path),
            'hasPassword': bool(proposal.price_password) if hasattr(proposal, 'price_password') else False
        }

    # Contracts
    contracts = []
    try:
        contracts = Contract.query.filter_by(
            consulting_project_id=id,
            document_type='contract'
        ).order_by(Contract.order_number).all()
    except Exception:
        pass

    if not contracts:
        try:
            all_contracts = Contract.query.filter_by(consulting_project_id=id).order_by(Contract.id).all()
            contracts = [c for c in all_contracts if not hasattr(c, 'document_type') or c.document_type in (None, 'contract', '')]
        except Exception:
            contracts = []

    result['contract']['items'] = [{
        'id': c.id,
        'orderNumber': getattr(c, 'order_number', None) or c.id,
        'fileName': getattr(c, 'file_name', None) or '',
        'description': getattr(c, 'description', None) or ''
    } for c in contracts]

    # Final reports
    try:
        final_reports = Contract.query.filter_by(
            consulting_project_id=id,
            document_type='final_report'
        ).all()
        result['completion']['finalReports'] = [{
            'contractId': r.id,
            'fileName': getattr(r, 'file_name', '') or ''
        } for r in final_reports]
    except Exception:
        result['completion']['finalReports'] = []

    return jsonify({'success': True, 'data': result})
