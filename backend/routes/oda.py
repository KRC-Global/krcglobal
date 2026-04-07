"""
GBMS - ODA Projects Routes
국제협력사업 프로젝트 관리 API
"""
from flask import Blueprint, request, jsonify, send_file
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from werkzeug.utils import secure_filename
from models import db, OdaProject, ActivityLog, ProfitabilityData, OdaManualData
from routes.auth import token_required, admin_required, permission_required
import os

oda_bp = Blueprint('oda', __name__)

# ODA 프로젝트 ID → WBS 코드 매핑 (profitability_data 연동)
ODA_PROJECT_WBS_MAPPING = {
    30: ['39102-001-07-0003'],  # 베트남 무병씨감자
    31: ['39102-001-07-0004'],  # 베트남 가축질병센터
    32: ['39102-001-07-0005'],  # 베트남 참깨 가치사슬
    33: ['39102-001-07-0006'],  # 캄보디아 고부가가치 채소생산
    34: ['39102-001-07-0007'],  # 캄보디아 농업비즈니스 및 농촌공동체
    35: ['39102-001-07-0008'],  # 캄보디아 AFSIS
    36: ['39102-001-07-0009'],  # 라오스 ITTC
    37: ['39102-001-07-0010'],  # 라오스 비엔티안주 디지털정보
    38: ['39102-001-07-0011'],  # 라오스 댐 안전관리
    39: ['39102-001-07-0012'],  # 필리핀 고품질 쌀 종자
    40: ['39102-001-07-0013'],  # 몽골 스마트농업단지
    41: ['39102-001-07-0014'],  # 베트남 홍강델타
    # 42: 라오스 쌍통군 사후관리는 profitability_data에 데이터 없음
    43: ['39102-001-07-0016'],  # 우즈베키스탄 첨단온실
    44: ['39102-001-07-0017'],  # 우즈베키스탄 씨감자
    45: ['39102-001-07-0018'],  # 우즈베키스탄 기후변화 플랫폼
    46: ['39102-001-07-0019'],  # 키르기스스탄 채소종자
    47: ['39102-001-07-0020'],  # 가나 라이스벨트
    48: ['39102-001-07-0021'],  # 탄자니아 잔지바르
    49: ['39102-001-07-0022'],  # 세네갈 중고 농기계
    50: ['39102-001-07-0023'],  # 세네갈 라이스벨트
    51: ['39102-001-07-0024'],  # 감비아 라이스벨트
    52: ['39102-001-07-0025'],  # 기니 라이스벨트
    53: ['39102-001-07-0026'],  # 우간다 라이스벨트
    54: ['39102-001-07-0027'],  # 카메룬 라이스벨트
    55: ['39102-001-07-0028'],  # 케냐 라이스벨트
    56: ['39102-001-07-0029'],  # 모잠비크 사후관리
    # 57: 캄보디아 농업용수는 profitability_data에 데이터 없음
    58: ['39102-001-07-0031'],  # 디지털 농업 역량강화(스리랑카/요르단)
}


def validate_project_data(data, is_update=False):
    """ODA 프로젝트 데이터 유효성 검증"""
    errors = []

    # 필수 필드 검증 (생성 시에만)
    if not is_update:
        if not data.get('title'):
            errors.append('사업명은 필수 입력 항목입니다.')
        if not data.get('country'):
            errors.append('국가는 필수 입력 항목입니다.')

    # 예산 검증
    if 'budget' in data and data['budget']:
        try:
            budget = float(data['budget'])
            if budget < 0:
                errors.append('예산은 0 이상이어야 합니다.')
        except (ValueError, TypeError):
            errors.append('예산은 숫자여야 합니다.')

    # 좌표 검증
    if 'latitude' in data and data['latitude']:
        try:
            lat = float(data['latitude'])
            if lat < -90 or lat > 90:
                errors.append('위도는 -90도에서 90도 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('위도는 숫자여야 합니다.')

    if 'longitude' in data and data['longitude']:
        try:
            lon = float(data['longitude'])
            if lon < -180 or lon > 180:
                errors.append('경도는 -180도에서 180도 사이여야 합니다.')
        except (ValueError, TypeError):
            errors.append('경도는 숫자여야 합니다.')

    return errors


@oda_bp.route('', methods=['GET'])
@token_required
def get_oda_projects(current_user):
    """Get all ODA projects with filters"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    country = request.args.get('country')
    continent = request.args.get('continent')
    status = request.args.get('status')
    project_type = request.args.get('project_type')
    search = request.args.get('search')

    query = OdaProject.query

    if country:
        query = query.filter(OdaProject.country == country)
    if continent:
        query = query.filter(OdaProject.continent == continent)
    if status:
        query = query.filter(OdaProject.status == status)
    if project_type:
        query = query.filter(OdaProject.project_type == project_type)
    if search:
        query = query.filter(
            db.or_(
                OdaProject.title.ilike(f'%{search}%'),
                OdaProject.description.ilike(f'%{search}%'),
                OdaProject.country.ilike(f'%{search}%')
            )
        )
        
    # Get year filter
    year = request.args.get('year', type=int)

    query = query.order_by(OdaProject.number.asc())
    
    # Get all matching projects before year filtering & pagination
    projects = query.all()
    
    # Year filtering (Python-side due to custom period format)
    if year:
        filtered_projects = []
        for project in projects:
            # Check period field: 'YY-'YY, 'YY.MM~'YY.MM, etc.
            try:
                if project.period:
                    # Clean string: remove quotes and spaces
                    clean_period = project.period.replace("'", "").replace("‘", "").replace("’", "").strip()
                    
                    # Split by tilde ~ or hyphen - (if it looks like a range separator)
                    import re
                    # Try splitting by ~ first as it is a common range separator
                    if '~' in clean_period:
                        parts = clean_period.split('~')
                    elif '-' in clean_period:
                        # Be careful with hyphen, it might be inside a date (e.g. 24-01)
                        # Heuristic: split by hyphen only if it separates two date-like parts
                        parts = clean_period.split('-')
                    else:
                        parts = [clean_period]
                    
                    start_year = None
                    end_year = None
                    
                    def parse_year_from_part(part):
                        if not part: return None
                        # Split by dot or hyphen to get year part
                        # e.g. 24.01 -> 24, 2024.01 -> 2024
                        subparts = re.split(r'[.]', part.strip())
                        if subparts and subparts[0].isdigit():
                            yy = int(subparts[0])
                            # Handle 2-digit years
                            if 0 <= yy <= 99:
                                return 1900 + yy if yy >= 50 else 2000 + yy
                            # Handle 4-digit years
                            if 1900 <= yy <= 2100:
                                return yy
                        return None

                    if len(parts) >= 2:
                        start_year = parse_year_from_part(parts[0])
                        end_year = parse_year_from_part(parts[1])
                    elif len(parts) == 1:
                        # Single date or period description
                        start_year = parse_year_from_part(parts[0])
                        # If single year provided, treat as start year or single year duration
                        # If we want to be strict, we might say start_year == year
                        
                    if start_year and end_year:
                        if start_year <= year <= end_year:
                            filtered_projects.append(project)
                    elif start_year and year >= start_year:
                         filtered_projects.append(project)
                    elif end_year and year <= end_year:
                        filtered_projects.append(project)
            except Exception:
                continue
        projects = filtered_projects

    # Manual Pagination
    total = len(projects)
    import math
    pages = math.ceil(total / per_page) if total > 0 else 1
    
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_items = projects[start_idx:end_idx]

    return jsonify({
        'success': True,
        'data': [project.to_dict() for project in paginated_items],
        'total': total,
        'pages': pages,
        'currentPage': page
    })


@oda_bp.route('/<int:project_id>', methods=['GET'])
@token_required
def get_oda_project(current_user, project_id):
    """Get single ODA project"""
    project = OdaProject.query.get_or_404(project_id)
    return jsonify({
        'success': True,
        'data': project.to_dict()
    })


@oda_bp.route('', methods=['POST'])
@token_required
def create_oda_project(current_user):
    """Create new ODA project"""
    data = request.get_json()

    if not data:
        return jsonify({
            'success': False,
            'message': '요청 데이터가 없습니다.'
        }), 400

    errors = validate_project_data(data, is_update=False)
    if errors:
        return jsonify({
            'success': False,
            'message': '입력 데이터 검증 실패',
            'errors': errors
        }), 400

    try:
        project = OdaProject(
            number=data.get('number'),
            country=data.get('country', '').strip(),
            latitude=data.get('latitude'),
            longitude=data.get('longitude'),
            title=data.get('title', '').strip(),
            title_en=data.get('titleEn', '').strip() if data.get('titleEn') else None,
            description=data.get('description', '').strip() if data.get('description') else None,
            contract_year=data.get('contractYear'),
            period=data.get('period'),
            budget=data.get('budget'),
            project_type=data.get('projectType', '').strip() if data.get('projectType') else None,
            status=data.get('status', '진행중'),
            continent=data.get('continent', '').strip() if data.get('continent') else None,
            client=data.get('client', '').strip() if data.get('client') else None,
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )

        db.session.add(project)
        db.session.flush()

        log = ActivityLog(
            user_id=current_user.id,
            action='create',
            entity_type='oda_project',
            entity_id=project.id,
            description=f'{current_user.name}님이 국제협력사업 프로젝트를 생성했습니다: {project.title}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '프로젝트가 생성되었습니다.',
            'data': project.to_dict()
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'프로젝트 생성 중 오류가 발생했습니다: {str(e)}'
        }), 500


@oda_bp.route('/<int:project_id>', methods=['PUT'])
@token_required
def update_oda_project(current_user, project_id):
    """Update ODA project"""
    project = OdaProject.query.get_or_404(project_id)
    data = request.get_json()

    if not data:
        return jsonify({
            'success': False,
            'message': '수정할 데이터가 없습니다.'
        }), 400

    errors = validate_project_data(data, is_update=True)
    if errors:
        return jsonify({
            'success': False,
            'message': '입력 데이터 검증 실패',
            'errors': errors
        }), 400

    try:
        if 'number' in data:
            project.number = data['number']
        if 'country' in data:
            project.country = data['country'].strip() if data['country'] else None
        if 'latitude' in data:
            project.latitude = data['latitude']
        if 'longitude' in data:
            project.longitude = data['longitude']
        if 'title' in data:
            project.title = data['title'].strip()
        if 'titleEn' in data:
            project.title_en = data['titleEn'].strip() if data['titleEn'] else None
        if 'description' in data:
            project.description = data['description'].strip() if data['description'] else None
        if 'contractYear' in data:
            project.contract_year = data['contractYear']
        if 'period' in data:
            project.period = data['period']
        if 'budget' in data:
            project.budget = data['budget']
        if 'projectType' in data:
            project.project_type = data['projectType'].strip() if data['projectType'] else None
        if 'status' in data:
            project.status = data['status']
        if 'continent' in data:
            project.continent = data['continent'].strip() if data['continent'] else None
        if 'client' in data:
            project.client = data['client'].strip() if data['client'] else None

        project.updated_at = datetime.utcnow()

        log = ActivityLog(
            user_id=current_user.id,
            action='update',
            entity_type='oda_project',
            entity_id=project.id,
            description=f'{current_user.name}님이 국제협력사업 프로젝트를 수정했습니다: {project.title}',
            ip_address=request.remote_addr
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': '프로젝트가 수정되었습니다.',
            'data': project.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'프로젝트 수정 중 오류가 발생했습니다: {str(e)}'
        }), 500


@oda_bp.route('/<int:project_id>', methods=['DELETE'])
@token_required
def delete_oda_project(current_user, project_id):
    """Delete ODA project"""
    project = OdaProject.query.get_or_404(project_id)
    project_title = project.title

    log = ActivityLog(
        user_id=current_user.id,
        action='delete',
        entity_type='oda_project',
        entity_id=project.id,
        description=f'{current_user.name}님이 국제협력사업 프로젝트를 삭제했습니다: {project_title}',
        ip_address=request.remote_addr
    )
    db.session.add(log)
    db.session.delete(project)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': '프로젝트가 삭제되었습니다.'
    })


@oda_bp.route('/export', methods=['GET'])
@token_required
def export_oda_projects(current_user):
    """Export ODA projects to Excel"""
    country = request.args.get('country')
    continent = request.args.get('continent')
    status = request.args.get('status')
    project_type = request.args.get('project_type')
    search = request.args.get('search')
    year = request.args.get('year', type=int)

    query = OdaProject.query

    if country:
        query = query.filter(OdaProject.country == country)
    if continent:
        query = query.filter(OdaProject.continent == continent)
    if status:
        query = query.filter(OdaProject.status == status)
    if project_type:
        query = query.filter(OdaProject.project_type == project_type)
    if search:
        query = query.filter(
            db.or_(
                OdaProject.title.ilike(f'%{search}%'),
                OdaProject.description.ilike(f'%{search}%'),
                OdaProject.country.ilike(f'%{search}%')
            )
        )

    query = query.order_by(OdaProject.number.asc())
    projects = query.all()

    # Year filtering (Python-side)
    if year:
        filtered_projects = []
        for project in projects:
            try:
                if project.period:
                    # Clean string: remove quotes and spaces
                    clean_period = project.period.replace("'", "").replace("‘", "").replace("’", "").strip()
                    
                    # Split by tilde ~ or hyphen - (if it looks like a range separator)
                    import re
                    # Try splitting by ~ first as it is a common range separator
                    if '~' in clean_period:
                        parts = clean_period.split('~')
                    elif '-' in clean_period:
                        # Be careful with hyphen, it might be inside a date (e.g. 24-01)
                        # Heuristic: split by hyphen only if it separates two date-like parts
                        parts = clean_period.split('-')
                    else:
                        parts = [clean_period]
                    
                    start_year = None
                    end_year = None
                    
                    def parse_year_from_part(part):
                        if not part: return None
                        # Split by dot or hyphen to get year part
                        # e.g. 24.01 -> 24, 2024.01 -> 2024
                        subparts = re.split(r'[.]', part.strip())
                        if subparts and subparts[0].isdigit():
                            yy = int(subparts[0])
                            # Handle 2-digit years
                            if 0 <= yy <= 99:
                                return 1900 + yy if yy >= 50 else 2000 + yy
                            # Handle 4-digit years
                            if 1900 <= yy <= 2100:
                                return yy
                        return None

                    if len(parts) >= 2:
                        start_year = parse_year_from_part(parts[0])
                        end_year = parse_year_from_part(parts[1])
                    elif len(parts) == 1:
                        # Single date or period description
                        start_year = parse_year_from_part(parts[0])
                        
                    if start_year and end_year:
                        if start_year <= year <= end_year:
                            filtered_projects.append(project)
                    elif start_year and year >= start_year:
                         filtered_projects.append(project)
                    elif end_year and year <= end_year:
                        filtered_projects.append(project)
            except Exception:
                continue
        projects = filtered_projects

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "국제협력사업"

    # Headers
    headers = [
        '번호', '사업명', '국가', '사업기간', '예산(백만원)', 
        '사업형태', '진행상태', '대륙', '위도', '경도'
    ]

    header_fill = PatternFill(start_color="0A3D62", end_color="0A3D62", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_num, project in enumerate(projects, 2):
        ws.cell(row=row_num, column=1, value=project.number)
        ws.cell(row=row_num, column=2, value=project.title)
        ws.cell(row=row_num, column=3, value=project.country)
        ws.cell(row=row_num, column=4, value=project.period)
        ws.cell(row=row_num, column=5, value=float(project.budget) if project.budget else None)
        ws.cell(row=row_num, column=6, value=project.project_type)
        ws.cell(row=row_num, column=7, value=project.status)
        ws.cell(row=row_num, column=8, value=project.continent)
        ws.cell(row=row_num, column=9, value=float(project.latitude) if project.latitude else None)
        ws.cell(row=row_num, column=10, value=float(project.longitude) if project.longitude else None)

    # Adjust widths
    column_widths = [8, 40, 15, 15, 15, 20, 12, 12, 12, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Log
    log = ActivityLog(
        user_id=current_user.id,
        action='export',
        entity_type='oda_project',
        description=f'{current_user.name}님이 국제협력사업 프로젝트 {len(projects)}건을 Excel로 다운로드했습니다.',
        ip_address=request.remote_addr
    )
    db.session.add(log)
    db.session.commit()

    filename = f"국제협력사업_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def parse_end_year_from_period(period):
    """period 문자열에서 종료 연도를 추출"""
    import re
    if not period:
        return None

    # Clean string: remove all types of quotes and spaces
    # Unicode quotes: ' (0x2018), ' (0x2019), ' (0x27), ` (0x60)
    clean_period = period.replace('\u2018', '').replace('\u2019', '').replace("'", "").replace('`', '').strip()

    # Split by tilde ~ or hyphen -
    if '~' in clean_period:
        parts = clean_period.split('~')
    elif '-' in clean_period:
        parts = clean_period.split('-')
    else:
        return None

    if len(parts) < 2:
        return None

    # Get the end part
    end_part = parts[-1].strip()

    # Extract year from end part (e.g., "24", "2024", "24.12")
    subparts = re.split(r'[.]', end_part)
    if subparts and subparts[0].isdigit():
        yy = int(subparts[0])
        # Handle 2-digit years
        if 0 <= yy <= 99:
            return 1900 + yy if yy >= 50 else 2000 + yy
        # Handle 4-digit years
        if 1900 <= yy <= 2100:
            return yy
    return None


@oda_bp.route('/stats', methods=['GET'])
@token_required
def get_oda_stats(current_user):
    """Get ODA projects statistics"""
    from datetime import datetime
    current_year = datetime.now().year

    # Get all projects
    all_projects = OdaProject.query.all()
    total = len(all_projects)

    # Calculate actual status based on period
    in_progress_count = 0
    completed_count = 0
    in_progress_budget = 0
    total_budget = 0

    for project in all_projects:
        budget = float(project.budget) if project.budget else 0
        total_budget += budget

        # Parse end year from period
        end_year = parse_end_year_from_period(project.period)

        # Determine actual status
        if end_year and end_year < current_year:
            # Period ended → completed (준공)
            completed_count += 1
        else:
            # Still in progress
            in_progress_count += 1
            in_progress_budget += budget

    # By continent
    continent_stats = db.session.query(
        OdaProject.continent,
        db.func.count(OdaProject.id)
    ).group_by(OdaProject.continent).all()

    # By country (top 10)
    country_stats = db.session.query(
        OdaProject.country,
        db.func.count(OdaProject.id)
    ).group_by(OdaProject.country).order_by(
        db.func.count(OdaProject.id).desc()
    ).limit(10).all()

    # By project type
    type_stats = db.session.query(
        OdaProject.project_type,
        db.func.count(OdaProject.id)
    ).group_by(OdaProject.project_type).all()

    # Projects with coordinates
    with_coords = OdaProject.query.filter(
        OdaProject.latitude.isnot(None),
        OdaProject.longitude.isnot(None)
    ).count()

    return jsonify({
        'success': True,
        'data': {
            'total': total,
            'inProgress': in_progress_count,
            'completed': completed_count,
            'byStatus': {
                '진행중': in_progress_count,
                '준공': completed_count
            },
            'byContinent': {continent or '미지정': count for continent, count in continent_stats},
            'byCountry': [{'country': country, 'count': count} for country, count in country_stats],
            'byType': {ptype or '미지정': count for ptype, count in type_stats},
            'totalBudget': float(total_budget),
            'inProgressBudget': float(in_progress_budget),
            'withCoordinates': with_coords,
            'coordinateRate': round(with_coords / total * 100, 1) if total > 0 else 0
        }
    })


@oda_bp.route('/countries', methods=['GET'])
@token_required
def get_oda_countries(current_user):
    """Get list of countries with projects"""
    countries = db.session.query(
        OdaProject.country,
        db.func.count(OdaProject.id).label('count')
    ).group_by(OdaProject.country).order_by(
        OdaProject.country.asc()
    ).all()

    return jsonify({
        'success': True,
        'data': [
            {'country': country, 'count': count}
            for country, count in countries
        ]
    })


@oda_bp.route('/continents', methods=['GET'])
@token_required
def get_oda_continents(current_user):
    """Get list of continents with projects"""
    continents = db.session.query(
        OdaProject.continent,
        db.func.count(OdaProject.id).label('count')
    ).group_by(OdaProject.continent).order_by(
        db.func.count(OdaProject.id).desc()
    ).all()

    return jsonify({
        'success': True,
        'data': [
            {'continent': continent or '미지정', 'count': count}
            for continent, count in continents
        ]
    })


@oda_bp.route('/current-year-projects', methods=['GET'])
@token_required
def get_current_year_oda_projects(current_user):
    """금년도 추진사업 목록 - profitability_data 연동"""
    try:
        current_year = datetime.now().year
        next_year = current_year + 1

        # Get all ODA projects
        all_projects = OdaProject.query.order_by(OdaProject.id.asc()).all()

        # Filter projects that are still in progress (end year >= current year)
        # 금년도 이후까지 진행되는 프로젝트만 표시
        in_progress_projects = []
        for project in all_projects:
            end_year = parse_end_year_from_period(project.period)
            if end_year is None or end_year >= current_year:
                in_progress_projects.append(project)

        if not in_progress_projects:
            return jsonify({
                'success': True,
                'currentYear': current_year,
                'nextYear': next_year,
                'yearColumns': [],
                'projects': [],
                'totals': {}
            })

        # Determine year columns (actual years with data only, no future years)
        all_years = set()
        project_data_list = []

        for project in in_progress_projects:
            # Parse start and end years
            start_year = None
            end_year = None

            if project.period:
                clean_period = project.period.replace('\u2018', '').replace('\u2019', '').replace("'", "").replace('`', '').strip()
                if '-' in clean_period:
                    parts = clean_period.split('-')
                    if len(parts) >= 2:
                        if parts[0].isdigit():
                            yy = int(parts[0])
                            start_year = 1900 + yy if yy >= 50 else 2000 + yy
                        if parts[1].isdigit():
                            yy = int(parts[1])
                            end_year = 1900 + yy if yy >= 50 else 2000 + yy

            # WBS 매핑을 통해 profitability_data에서 수익/비용 가져오기
            wbs_codes = ODA_PROJECT_WBS_MAPPING.get(project.id, [])
            revenues_by_year = {}
            costs_by_year = {}
            total_past_revenue = 0
            total_past_cost = 0

            # 1. profitability_data에서 데이터 가져오기 (2024년 이후)
            if wbs_codes:
                # 모든 WBS 코드에 대해 profitability_data 조회
                profitability_records = ProfitabilityData.query.filter(
                    ProfitabilityData.wbs_code.in_(wbs_codes),
                    ProfitabilityData.year <= current_year  # 금년도 포함
                ).all()

                for record in profitability_records:
                    year = record.year
                    revenue = int(record.revenue or 0)  # 천원 단위
                    # 비용 = 직접비 + 인건비 + 경비
                    direct_cost = int(record.direct_cost or 0)
                    labor_cost = int(record.labor_cost or 0)
                    expense = int(record.expense or 0)
                    total_cost = direct_cost + labor_cost + expense

                    all_years.add(year)

                    if year not in revenues_by_year:
                        revenues_by_year[year] = 0
                    revenues_by_year[year] += revenue
                    total_past_revenue += revenue

                    if year not in costs_by_year:
                        costs_by_year[year] = 0
                    costs_by_year[year] += total_cost
                    total_past_cost += total_cost

            # 2. 수기 입력 데이터 가져오기 (2024년 이전 데이터)
            manual_records = OdaManualData.query.filter_by(oda_project_id=project.id).all()
            for record in manual_records:
                year = record.year
                revenue = record.revenue or 0  # 이미 천원 단위
                cost = record.cost or 0

                all_years.add(year)

                if year not in revenues_by_year:
                    revenues_by_year[year] = 0
                revenues_by_year[year] += revenue
                total_past_revenue += revenue

                if year not in costs_by_year:
                    costs_by_year[year] = 0
                costs_by_year[year] += cost
                total_past_cost += cost

            # 총 예산 (백만원)
            budget = float(project.budget) if project.budget else 0
            budget_in_thousands = budget * 1000  # 백만원 → 천원

            # 내년도 이후 수익 = 총예산 - 그간 수익 합계
            future_revenue = max(0, budget_in_thousands - total_past_revenue)

            # 내년도 이후 비용 = 총예산 - 그간 비용 합계 (추정)
            future_cost = max(0, budget_in_thousands - total_past_cost)

            project_data_list.append({
                'id': project.id,
                'titleKr': project.title,
                'country': project.country,
                'period': project.period,
                'budget': budget,
                'budgetInThousands': budget_in_thousands,
                'startYear': start_year,
                'endYear': end_year,
                'revenues': revenues_by_year,
                'costs': costs_by_year,
                'futureRevenue': round(future_revenue),
                'futureCost': round(future_cost),
                'wbsCodes': wbs_codes
            })

        # Year columns: 2020년부터 금년도까지 모든 연도 포함
        year_columns = list(range(2020, current_year + 1))

        # 연도별 합계 계산
        totals_by_year = {y: 0 for y in year_columns}
        costs_by_year_totals = {y: 0 for y in year_columns}
        total_future = 0
        total_future_cost = 0
        total_budget_sum = 0

        for project_data in project_data_list:
            for year, revenue in project_data['revenues'].items():
                if year in totals_by_year:
                    totals_by_year[year] += revenue
            for year, cost in project_data['costs'].items():
                if year in costs_by_year_totals:
                    costs_by_year_totals[year] += cost
            total_future += project_data['futureRevenue']
            total_future_cost += project_data['futureCost']
            total_budget_sum += project_data['budgetInThousands']

        return jsonify({
            'success': True,
            'currentYear': current_year,
            'nextYear': next_year,
            'yearColumns': year_columns,
            'projects': project_data_list,
            'totals': {
                'totalBudget': round(total_budget_sum),
                'revenues': {year: round(totals_by_year[year]) for year in year_columns},
                'costs': {year: round(costs_by_year_totals[year]) for year in year_columns},
                'futureRevenue': round(total_future),
                'futureCost': round(total_future_cost)
            }
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'금년도 추진사업 목록 조회 중 오류가 발생했습니다: {str(e)}'
        }), 500


# ==================== 수기 입력 데이터 관리 (2024년 이전) ====================

@oda_bp.route('/manual-data/<int:oda_project_id>', methods=['GET'])
@token_required
def get_manual_data(current_user, oda_project_id):
    """특정 프로젝트의 수기 입력 데이터 조회"""
    try:
        records = OdaManualData.query.filter_by(oda_project_id=oda_project_id).order_by(OdaManualData.year.asc()).all()
        return jsonify({
            'success': True,
            'data': [record.to_dict() for record in records]
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'수기 입력 데이터 조회 중 오류가 발생했습니다: {str(e)}'
        }), 500


@oda_bp.route('/manual-data', methods=['POST'])
@permission_required('oda')
def save_manual_data(current_user):
    """수기 입력 데이터 저장 (생성 또는 수정) - ODA 권한 필요"""
    try:
        data = request.get_json()
        oda_project_id = data.get('odaProjectId')
        year = data.get('year')
        revenue = data.get('revenue', 0)
        cost = data.get('cost', 0)

        if not oda_project_id or not year:
            return jsonify({
                'success': False,
                'message': '프로젝트 ID와 연도는 필수입니다.'
            }), 400

        # 프로젝트 존재 여부 확인
        project = OdaProject.query.get(oda_project_id)
        if not project:
            return jsonify({
                'success': False,
                'message': '해당 프로젝트를 찾을 수 없습니다.'
            }), 404

        # 기존 레코드 확인 (upsert 로직)
        existing = OdaManualData.query.filter_by(
            oda_project_id=oda_project_id,
            year=year
        ).first()

        if existing:
            # 수정
            existing.revenue = revenue
            existing.cost = cost
            existing.updated_by = current_user.id
            existing.updated_at = datetime.utcnow()
            message = f'{year}년 데이터가 수정되었습니다.'
        else:
            # 생성
            new_record = OdaManualData(
                oda_project_id=oda_project_id,
                year=year,
                revenue=revenue,
                cost=cost,
                created_by=current_user.id,
                updated_by=current_user.id
            )
            db.session.add(new_record)
            message = f'{year}년 데이터가 추가되었습니다.'

        db.session.commit()

        # Activity log
        log = ActivityLog(
            user_id=current_user.id,
            action='update' if existing else 'create',
            entity_type='oda_manual_data',
            entity_id=existing.id if existing else new_record.id,
            description=f'{project.title} - {message}'
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': message
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'수기 입력 데이터 저장 중 오류가 발생했습니다: {str(e)}'
        }), 500


@oda_bp.route('/manual-data/<int:id>', methods=['DELETE'])
@permission_required('oda')
def delete_manual_data(current_user, id):
    """수기 입력 데이터 삭제 - ODA 권한 필요"""
    try:
        record = OdaManualData.query.get(id)
        if not record:
            return jsonify({
                'success': False,
                'message': '해당 데이터를 찾을 수 없습니다.'
            }), 404

        project = OdaProject.query.get(record.oda_project_id)
        project_title = project.title if project else '알 수 없음'
        year = record.year

        db.session.delete(record)
        db.session.commit()

        # Activity log
        log = ActivityLog(
            user_id=current_user.id,
            action='delete',
            entity_type='oda_manual_data',
            entity_id=id,
            description=f'{project_title} - {year}년 수기 입력 데이터 삭제'
        )
        db.session.add(log)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{year}년 데이터가 삭제되었습니다.'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'수기 입력 데이터 삭제 중 오류가 발생했습니다: {str(e)}'
        }), 500
